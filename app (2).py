import csv
import json
import math
import os
import re
import ssl
import textwrap
import urllib.error
import urllib.parse
import urllib.request
from io import BytesIO
from pathlib import Path
from typing import List

import certifi
import numpy as np
import openai
import pandas as pd
pd.set_option('mode.string_storage', 'python')
import streamlit as st
import networkx as nx
from dotenv import load_dotenv
try:
    import plotly.graph_objects as go
    HAS_PLOTLY = True
except ImportError:
    HAS_PLOTLY = False

try:
    import ollama
    HAS_OLLAMA = True
except ImportError:
    HAS_OLLAMA = False

try:
    from streamlit_agraph import Config, Edge, Node, agraph
    HAS_AGRAPH = True
except ImportError:
    HAS_AGRAPH = False

from transaction_networks import AnalyzerConfig, FinancialCrimeNetworkAnalyzer, SCHEMA_COLUMNS

from dotenv import load_dotenv

load_dotenv() # Works locally

OPENAI_API_KEY = (
st.secrets.get("OPENAI_API_KEY")
or os.getenv("OPENAI_API_KEY")
or os.getenv("openai_api_key")
)

GEMINI_API_KEY = (
st.secrets.get("GEMINI_API_KEY")
or st.secrets.get("GOOGLE_API_KEY")
or os.getenv("GEMINI_API_KEY")
or os.getenv("GOOGLE_API_KEY")
)

GROQ_API_KEY = (
st.secrets.get("GROQ_API_KEY")
or os.getenv("GROQ_API_KEY")
)
if OPENAI_API_KEY:
    openai.api_key = OPENAI_API_KEY

st.set_page_config(page_title="Atlas Detection Lab", layout="wide", page_icon="🔍")

st.markdown("""
<style>
/* Theme tokens derived from UI_Theme.jpg */
:root {
    --bg-0: #020817;
    --bg-1: #06162d;
    --bg-2: #0c2143;
    --panel: #0a1935;
    --panel-2: #11294f;
    --ink: #dbe8ff;
    --muted: #9db6dd;
    --line: #224a84;
    --accent: #0050a0;
    --accent-2: #0060b0;
    --accent-soft: #003070;
}

/* Global canvas */
[data-testid="stAppViewContainer"] {
    background: radial-gradient(1200px 700px at 86% -10%, #113469 0%, transparent 55%),
                            radial-gradient(900px 600px at -10% 10%, #072349 0%, transparent 48%),
                            linear-gradient(160deg, var(--bg-0) 0%, var(--bg-1) 45%, var(--bg-2) 100%);
}
[data-testid="stHeader"] {
    background: rgba(2, 8, 23, 0.38) !important;
    border-bottom: 1px solid rgba(88, 143, 220, 0.2);
    backdrop-filter: blur(6px);
}

/* Sidebar shell */
[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #061123 0%, #0a1d3d 100%) !important;
    border-right: 1px solid rgba(75, 129, 205, 0.32);
}

/* Sidebar form controls */
[data-testid="stSidebar"] input,
[data-testid="stSidebar"] textarea,
[data-testid="stSidebar"] [data-baseweb="select"] [data-testid="stMarkdownContainer"] p,
[data-testid="stSidebar"] [data-baseweb="select"] span,
[data-testid="stSidebar"] [data-baseweb="select"] div[class*="ValueContainer"] *,
[data-testid="stSidebar"] [data-baseweb="select"] div[class*="singleValue"],
[data-testid="stSidebar"] [data-baseweb="select"] div[class*="placeholder"],
[data-testid="stSidebar"] [data-baseweb="input"] input {
    color: var(--ink) !important;
    background: #0b2142 !important;
    font-weight: 500 !important;
}
[data-testid="stSidebar"] input,
[data-testid="stSidebar"] textarea,
[data-testid="stSidebar"] [data-baseweb="select"] > div {
    border: 1px solid var(--line) !important;
    border-radius: 8px !important;
    box-shadow: inset 0 0 0 1px rgba(0, 96, 176, 0.15);
}
[data-testid="stSidebar"] [data-baseweb="select"] div[class*="Option"] {
    color: #d7e7ff !important;
    background: #0a1a36 !important;
}
/* Keep sidebar selectboxes visually static while preserving normal typing
    behavior for main-page selectors such as the customer picker. */
[data-testid="stSidebar"] [data-baseweb="select"] input,
[data-testid="stSidebar"] [data-baseweb="select"] input[type="text"],
[data-testid="stSidebar"] div[data-baseweb="select"] input {
    caret-color: transparent !important;
    color: transparent !important;
    text-shadow: none !important;
    pointer-events: none !important;
    width: 0 !important;
    min-width: 0 !important;
    max-width: 0 !important;
    opacity: 0 !important;
}
[data-testid="stSidebar"] [data-baseweb="popover"],
[data-testid="stSidebar"] [role="listbox"],
[data-testid="stSidebar"] [role="option"] {
    background: #0a1a36 !important;
    color: #d7e7ff !important;
}
[data-testid="stSidebar"] [role="option"]:hover {
    background: #11335f !important;
}
[data-testid="stSidebar"] [data-testid="stNumberInputStepDown"],
[data-testid="stSidebar"] [data-testid="stNumberInputStepUp"],
[data-testid="stSidebar"] button[data-testid="stNumberInputStepDown"],
[data-testid="stSidebar"] button[data-testid="stNumberInputStepUp"],
[data-testid="stSidebar"] .stNumberInput button {
    color: #eef5ff !important;
    background: linear-gradient(180deg, var(--accent) 0%, var(--accent-2) 100%) !important;
    border: 1px solid #2a65a6 !important;
    border-radius: 5px !important;
    font-weight: 700 !important;
}
[data-testid="stSidebar"] [data-testid="stNumberInputStepDown"]:hover,
[data-testid="stSidebar"] [data-testid="stNumberInputStepUp"]:hover,
[data-testid="stSidebar"] button[data-testid="stNumberInputStepDown"]:hover,
[data-testid="stSidebar"] button[data-testid="stNumberInputStepUp"]:hover,
[data-testid="stSidebar"] .stNumberInput button:hover {
    filter: brightness(1.12);
}
[data-testid="stSidebar"] .stNumberInput input::placeholder { color: #8eaad0 !important; }
[data-testid="stSidebar"] .stSelectbox label,
[data-testid="stSidebar"] .stNumberInput label,
[data-testid="stSidebar"] .stCheckbox label {
    color: #d3e4ff !important;
    font-size: 0.82rem;
    font-weight: 600 !important;
}
[data-testid="stSidebar"] h2,
[data-testid="stSidebar"] h3 {
    color: #e7f1ff !important;
}

/* App cards and panels */
[data-testid="stVerticalBlockBorderWrapper"] {
    background: linear-gradient(160deg, rgba(8, 27, 56, 0.85) 0%, rgba(8, 22, 44, 0.85) 100%);
    border: 1px solid rgba(63, 112, 183, 0.34);
    border-radius: 12px;
    box-shadow: 0 12px 28px rgba(0, 0, 0, 0.25), inset 0 1px 0 rgba(165, 206, 255, 0.06);
}

/* Header */
.app-header {
    background: linear-gradient(130deg, #07162f 0%, #0f2d56 58%, #0050a0 100%);
    border: 1px solid rgba(106, 157, 229, 0.34);
    border-radius: 14px;
    padding: 1.4rem 2rem 1.2rem;
    margin-bottom: 1.2rem;
    box-shadow: 0 14px 32px rgba(0, 0, 0, 0.28);
}
.app-header h1 {
    color: #f2f7ff;
    margin: 0 0 0.25rem;
    font-size: 1.62rem;
    letter-spacing: 0.01em;
}
.app-header p {
    color: #b7d0f0;
    margin: 0;
    font-size: 0.88rem;
}

/* Section title and text */
.card-title {
    font-size: 1.05rem;
    font-weight: 700;
    color: #dceaff;
    margin-bottom: 0.6rem;
    padding-bottom: 0.45rem;
    border-bottom: 2px solid rgba(56, 106, 176, 0.56);
    letter-spacing: 0.01em;
}
[data-testid="stMarkdownContainer"] p,
[data-testid="stCaptionContainer"] {
    color: #c1d7f7;
}

/* Keep Streamlit metric labels readable in narrow columns */
[data-testid="stMetricLabel"] p {
    white-space: normal !important;
    line-height: 1.2 !important;
}

/* Risk badge */
.badge {
    display: inline-block;
    padding: 2px 10px;
    border-radius: 12px;
    font-size: 0.78rem;
    font-weight: 700;
    margin-right: 4px;
}
.badge-red { background: #3f1014; color: #ff9aa0; }
.badge-orange { background: #44280c; color: #ffcb85; }
.badge-purple { background: #2e1d49; color: #ccb7ff; }
.badge-blue { background: #0f3157; color: #8fd0ff; }

/* AI summary and actions */
.genai-card {
    background: linear-gradient(140deg, #0a1d3d 0%, #0c2749 100%);
    border-left: 4px solid var(--accent-2);
    border: 1px solid rgba(90, 145, 215, 0.45);
    border-radius: 10px;
    padding: 1rem 1.2rem;
    font-size: 0.94rem;
    line-height: 1.65;
    letter-spacing: 0.005em;
    color: #f0f6ff;
    box-shadow: inset 0 1px 0 rgba(150, 195, 255, 0.08);
}
.action-item {
    background: linear-gradient(160deg, rgba(11, 35, 70, 0.85) 0%, rgba(10, 28, 54, 0.85) 100%);
    border: 1px solid rgba(59, 112, 184, 0.34);
    border-radius: 8px;
    padding: 0.6rem 1rem;
    margin-bottom: 0.45rem;
    font-size: 0.9rem;
    color: #d9e9ff;
}

/* Buttons */
.stDownloadButton > button,
[data-testid="stButton"] > button {
    width: 100%;
    border-radius: 8px !important;
    background: linear-gradient(180deg, var(--accent) 0%, var(--accent-2) 100%) !important;
    color: #f0f7ff !important;
    border: 1px solid #2f6cb0 !important;
    font-weight: 700 !important;
    box-shadow: 0 8px 20px rgba(0, 48, 112, 0.35);
}
.stDownloadButton > button:hover,
[data-testid="stButton"] > button:hover {
    filter: brightness(1.1);
    transform: translateY(-1px);
}
/* ===========================
Metric Values (FIX)
=========================== */

[data-testid="stMetricValue"] {
color: #FFFFFF !important;
opacity: 1 !important;
font-weight: 700 !important;
}

[data-testid="stMetricValue"] > div {
color: #FFFFFF !important;
opacity: 1 !important;
}

[data-testid="stMetricValue"] * {
color: #FFFFFF !important;
opacity: 1 !important;
-webkit-text-fill-color: #FFFFFF !important;
}

[data-testid="stMetricLabel"] {
color: #CFE5FF !important;
font-weight: 600 !important;
}

[data-testid="stMetricDelta"] {
color: #FFFFFF !important;
opacity: 1 !important;
}
</style>
""", unsafe_allow_html=True)


def card_title(icon: str, text: str) -> None:
    st.markdown(f'<div class="card-title">{icon}&nbsp; {text}</div>', unsafe_allow_html=True)


def to_csv_bytes(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode("utf-8")


def _normalize_pdf_text(text: str) -> str:
    value = "" if text is None else str(text)
    value = re.sub(r"<br\s*/?>", "\n", value, flags=re.IGNORECASE)
    value = re.sub(r"</?[^>]+>", "", value)
    replacements = {
        "•": "-",
        "–": "-",
        "—": "-",
        "↗": "->",
        "→": "->",
        "≤": "<=",
        "≥": ">=",
        "🔴": "",
        "🟠": "",
        "🟡": "",
        "🟢": "",
        "🤖": "",
        "🧭": "",
        "📊": "",
        "⚠️": "",
        "🚨": "",
        "⬇️": "",
        "📥": "",
    }
    for old, new in replacements.items():
        value = value.replace(old, new)
    value = value.replace("\t", "    ").replace("\r\n", "\n").replace("\r", "\n")
    return value.encode("latin-1", "replace").decode("latin-1")


def _escape_pdf_text(text: str) -> str:
    return _normalize_pdf_text(text).replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def _wrap_pdf_lines(text: str, width: int = 95) -> List[str]:
    normalized = _normalize_pdf_text(text)
    lines: List[str] = []
    for paragraph in normalized.split("\n"):
        stripped = paragraph.strip()
        if not stripped:
            lines.append("")
            continue
        wrapped = textwrap.wrap(stripped, width=width, break_long_words=False, break_on_hyphens=False)
        lines.extend(wrapped or [stripped])
    return lines


def _dataframe_pdf_elements(title: str, df: pd.DataFrame, max_rows: int | None = None) -> list[tuple[str, str]]:
    elements: list[tuple[str, str]] = [("section", title)]
    if df is None or df.empty:
        elements.append(("body", "No rows available."))
        return elements

    display_df = df.copy().fillna("")
    if max_rows is not None and len(display_df) > max_rows:
        display_df = display_df.head(max_rows).copy()
        truncated = True
    else:
        truncated = False

    for idx, row in enumerate(display_df.to_dict(orient="records"), start=1):
        parts = []
        for col, value in row.items():
            if isinstance(value, float):
                value_str = f"{value:.2f}"
            else:
                value_str = str(value)
            if value_str and value_str.lower() != "nan":
                parts.append(f"{col}: {value_str}")
        elements.append(("body", f"{idx}. " + " | ".join(parts)))

    if truncated:
        elements.append(("body", f"Showing first {len(display_df)} of {len(df)} rows in this PDF section."))
    return elements


def _pdf_elements_to_bytes(elements: list[tuple[str, str]]) -> bytes:
    page_width = 612
    page_height = 792
    margin_left = 44
    margin_top = 748
    margin_bottom = 48
    styles = {
        "title": {"size": 16, "leading": 22, "width": 72, "gap": 6},
        "section": {"size": 12, "leading": 16, "width": 88, "gap": 4},
        "meta": {"size": 10, "leading": 13, "width": 96, "gap": 2},
        "body": {"size": 10, "leading": 13, "width": 96, "gap": 2},
    }

    pages: list[list[tuple[str, str, int, float]]] = []
    current_page: list[tuple[str, str, int, float]] = []
    y = float(margin_top)

    for style, text in elements:
        spec = styles.get(style, styles["body"])
        wrapped_lines = _wrap_pdf_lines(text, width=int(spec["width"]))
        if not wrapped_lines:
            wrapped_lines = [""]

        for line in wrapped_lines:
            if y < margin_bottom + spec["leading"]:
                pages.append(current_page)
                current_page = []
                y = float(margin_top)
            current_page.append((style, line, int(spec["size"]), y))
            y -= float(spec["leading"])
        y -= float(spec["gap"])

    if current_page:
        pages.append(current_page)
    if not pages:
        pages = [[("body", "No content available.", 10, float(margin_top))]]

    objects: dict[int, bytes] = {}
    objects[1] = b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>"

    page_object_ids = []
    next_obj_id = 3
    for page in pages:
        content_lines = []
        for style, line, font_size, line_y in page:
            escaped = _escape_pdf_text(line)
            content_lines.append(f"BT /F1 {font_size} Tf {margin_left} {line_y:.2f} Td ({escaped}) Tj ET")
        content_stream = "\n".join(content_lines).encode("latin-1", "replace")
        content_obj_id = next_obj_id
        page_obj_id = next_obj_id + 1
        next_obj_id += 2
        objects[content_obj_id] = (
            f"<< /Length {len(content_stream)} >>\nstream\n".encode("latin-1")
            + content_stream
            + b"\nendstream"
        )
        objects[page_obj_id] = (
            f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 {page_width} {page_height}] ".encode("latin-1")
            + f"/Resources << /Font << /F1 1 0 R >> >> /Contents {content_obj_id} 0 R >>".encode("latin-1")
        )
        page_object_ids.append(page_obj_id)

    kids = " ".join([f"{obj_id} 0 R" for obj_id in page_object_ids])
    objects[2] = f"<< /Type /Pages /Kids [{kids}] /Count {len(page_object_ids)} >>".encode("latin-1")
    catalog_obj_id = next_obj_id
    objects[catalog_obj_id] = b"<< /Type /Catalog /Pages 2 0 R >>"

    pdf = bytearray()
    pdf.extend(b"%PDF-1.4\n%\xe2\xe3\xcf\xd3\n")
    offsets = [0]
    for obj_id in sorted(objects.keys()):
        offsets.append(len(pdf))
        pdf.extend(f"{obj_id} 0 obj\n".encode("latin-1"))
        pdf.extend(objects[obj_id])
        pdf.extend(b"\nendobj\n")

    xref_offset = len(pdf)
    pdf.extend(f"xref\n0 {len(offsets)}\n".encode("latin-1"))
    pdf.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        pdf.extend(f"{offset:010d} 00000 n \n".encode("latin-1"))
    pdf.extend(
        f"trailer\n<< /Size {len(offsets)} /Root {catalog_obj_id} 0 R >>\nstartxref\n{xref_offset}\n%%EOF".encode("latin-1")
    )
    return bytes(pdf)


def build_network_summary_pdf_bytes(
    analyzer,
    target_customer_id: str,
    selected_network: str,
    selected_node: str,
    network_summary_df: pd.DataFrame,
    top_nodes_df: pd.DataFrame,
    theme_table_df: pd.DataFrame,
    target_summary_text: str,
    target_summary_status: str,
    target_actions: List[str],
    selected_node_summary_text: str,
    selected_node_summary_status: str,
    selected_node_actions: List[str],
    recommendation_text: str,
    recommendation_status: str,
) -> bytes:
    elements: list[tuple[str, str]] = [
        ("title", "Atlas Detection Lab - Network Summary Export"),
        ("meta", f"Target customer: {target_customer_id}"),
        ("meta", f"Selected network focus: {selected_network}"),
        ("meta", f"Selected node: {selected_node}"),
        ("meta", f"Analysis window: {getattr(analyzer, 'window_label', '')}"),
        ("meta", "This PDF contains exported network-level results and the current AI or deterministic narratives shown in the UI."),
    ]
    elements.extend(_dataframe_pdf_elements("Network Summary Ranking", network_summary_df))
    elements.extend(_dataframe_pdf_elements("Top Risky Nodes", top_nodes_df, max_rows=40))
    elements.extend(_dataframe_pdf_elements("Theme / Subtheme Trigger Log", theme_table_df, max_rows=80))

    elements.append(("section", "Target Customer Investigator Summary"))
    if target_summary_status:
        elements.append(("meta", f"Status: {target_summary_status}"))
    elements.append(("body", target_summary_text or "No target customer summary available."))

    elements.append(("section", "Selected Node Risk Summary"))
    if selected_node_summary_status:
        elements.append(("meta", f"Status: {selected_node_summary_status}"))
    elements.append(("body", selected_node_summary_text or "No selected-node summary available."))

    elements.append(("section", "Network Recommendation"))
    if recommendation_status:
        elements.append(("meta", f"Status: {recommendation_status}"))
    elements.append(("body", recommendation_text or "No recommendation available."))

    return _pdf_elements_to_bytes(elements)


def _autosize_excel_sheet(ws, min_width: int = 14, max_width: int = 60) -> None:
    for column_cells in ws.columns:
        column_letter = column_cells[0].column_letter
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[column_letter].width = max(min_width, min(max_width, max_len + 2))


def _write_excel_dataframe(writer, sheet_name: str, df: pd.DataFrame) -> None:
    display_df = df.copy() if isinstance(df, pd.DataFrame) else pd.DataFrame()
    display_df.to_excel(writer, sheet_name=sheet_name, index=False)
    ws = writer.book[sheet_name]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _write_excel_narrative_sheet(
    writer,
    sheet_name: str,
    title: str,
    status: str,
    summary_text: str,
    actions: List[str],
    include_actions: bool = True,
) -> None:
    rows = []
    rows.append({"Section": "Title", "Content": title})
    if status:
        rows.append({"Section": "Status", "Content": status})
    rows.append({"Section": "Summary", "Content": summary_text or "No summary available."})
    if include_actions:
        if actions:
            for idx, action in enumerate(actions, start=1):
                rows.append({"Section": f"Action {idx}", "Content": action})
        else:
            rows.append({"Section": "Actions", "Content": "No actions available."})
    pd.DataFrame(rows).to_excel(writer, sheet_name=sheet_name, index=False)


def build_network_summary_excel_bytes(
    analyzer,
    target_customer_id: str,
    selected_network: str,
    selected_node: str,
    network_summary_df: pd.DataFrame,
    top_nodes_df: pd.DataFrame,
    theme_table_df: pd.DataFrame,
    all_networks_tx_df: pd.DataFrame,
    target_summary_text: str,
    target_summary_status: str,
    target_actions: List[str],
    selected_node_summary_text: str,
    selected_node_summary_status: str,
    selected_node_actions: List[str],
    recommendation_text: str,
    recommendation_status: str,
) -> bytes:
    from openpyxl.styles import Alignment, Font, PatternFill

    overview_rows = [
        {"Field": "Export Title", "Value": "Atlas Detection Lab - Network Summary Export"},
        {"Field": "Target Customer", "Value": str(target_customer_id)},
        {"Field": "Selected Network Focus", "Value": str(selected_network)},
        {"Field": "Selected Node", "Value": str(selected_node)},
        {"Field": "Analysis Window", "Value": str(getattr(analyzer, "window_label", ""))},
        {
            "Field": "Contents",
            "Value": "Network summary, risky nodes, theme triggers, all-network transactions, target summary, selected node summary, and network recommendation.",
        },
    ]

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        pd.DataFrame(overview_rows).to_excel(writer, sheet_name="Overview", index=False)
        _write_excel_dataframe(writer, "Network Summary", network_summary_df)
        _write_excel_dataframe(writer, "Top Risky Nodes", top_nodes_df)
        _write_excel_dataframe(writer, "Theme Triggers", theme_table_df)
        _write_excel_dataframe(writer, "All Transactions", all_networks_tx_df)
        _write_excel_narrative_sheet(
            writer,
            sheet_name="Target Summary",
            title=f"Target Customer Summary - {target_customer_id}",
            status=target_summary_status,
            summary_text=target_summary_text,
            actions=target_actions,
            include_actions=False,
        )
        _write_excel_narrative_sheet(
            writer,
            sheet_name="Selected Node Summary",
            title=f"Selected Node Summary - {selected_node}",
            status=selected_node_summary_status,
            summary_text=selected_node_summary_text,
            actions=selected_node_actions,
            include_actions=False,
        )
        _write_excel_narrative_sheet(
            writer,
            sheet_name="Recommendation",
            title="Network Recommendation",
            status=recommendation_status,
            summary_text=recommendation_text,
            actions=[],
            include_actions=False,
        )

        workbook = writer.book
        header_fill = PatternFill(fill_type="solid", fgColor="0F2D56")
        header_font = Font(color="FFFFFF", bold=True)
        title_fill = PatternFill(fill_type="solid", fgColor="0050A0")
        title_font = Font(color="FFFFFF", bold=True)

        for ws in workbook.worksheets:
            ws.sheet_view.showGridLines = True
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

            if ws.title in {"Target Summary", "Selected Node Summary", "Recommendation"}:
                ws["A2"].fill = title_fill
                ws["A2"].font = title_font
                ws["B2"].fill = title_fill
                ws["B2"].font = title_font
                ws.column_dimensions["A"].width = 20
                ws.column_dimensions["B"].width = 90
            elif ws.title == "Overview":
                ws.column_dimensions["A"].width = 24
                ws.column_dimensions["B"].width = 90
            else:
                _autosize_excel_sheet(ws)

        writer.book.active = 0

    return buffer.getvalue()


def _format_action_lines(text: str) -> List[str]:
    if not text:
        return []
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    cleaned = [re.sub(r'^\s*\d+[\)\.\-\s]*', '', line).strip() for line in lines]
    return cleaned
def _trim_incomplete_tail(text: str) -> str:
    if not text:
        return ""
    cleaned = str(text).strip()
    if not cleaned:
        return ""
    if cleaned[-1] in ".!?":
        return cleaned
    last_sentence_break = max(cleaned.rfind(". "), cleaned.rfind("! "), cleaned.rfind("? "))
    if last_sentence_break != -1:
        return cleaned[: last_sentence_break + 1].strip()
    last_line_break = cleaned.rfind("\n")
    if last_line_break != -1:
        return cleaned[:last_line_break].strip()
    return cleaned + "."

def _openai_error_status(ex: Exception) -> str:
    message = str(ex)
    code = getattr(ex, 'code', None)
    status = getattr(ex, 'status_code', None) or getattr(ex, 'http_status', None)
    if hasattr(ex, 'error'):
        err = getattr(ex, 'error')
        if isinstance(err, dict):
            message = err.get('message', message)
            code = err.get('code', code)
    is_quota = status == 429 or code == 'insufficient_quota' or 'insufficient_quota' in message or 'quota' in message.lower()
    if is_quota:
        if GEMINI_API_KEY:
            return f"OpenAI quota exceeded ({message}); Gemini is configured, so switch the AI provider to Gemini in the sidebar or keep deterministic fallback."
        return f"OpenAI quota exceeded ({message}); set GEMINI_API_KEY and select Gemini or use deterministic fallback."
    return f"OpenAI unavailable ({message}); showing deterministic AI Summary Generator output."


def _openai_chat(prompt: str, model: str = "gpt-5.5", temperature: float = 0.2, max_tokens: int = 1200) -> str:
    if not openai.api_key:
        raise RuntimeError("OpenAI API key is not configured. Set OPENAI_API_KEY in .env")

    # GPT-5.x reasoning models reject a custom 'temperature' and the legacy 'max_tokens'
    # field on Chat Completions; they require 'max_completion_tokens' instead.
    is_reasoning_model = model.startswith("gpt-5")
    kwargs = {}
    if is_reasoning_model:
        kwargs["max_completion_tokens"] = max_tokens
    else:
        kwargs["max_tokens"] = max_tokens
        kwargs["temperature"] = temperature

    response = openai.chat.completions.create(
        model=model,
        messages=[
            {"role": "system", "content": "You are a concise and factual financial crime investigation assistant."},
            {"role": "user", "content": prompt},
        ],
        **kwargs,
    )
    return str(response.choices[0].message.content).strip()


def _gemini_chat(prompt: str, model: str = "gemini-2.5-flash", temperature: float = 0.2, max_tokens: int = 600) -> str:
    if not GEMINI_API_KEY:
        raise RuntimeError("Gemini API key is not configured. Set GEMINI_API_KEY or GOOGLE_API_KEY in .env")

    fallback_order = ["gemini-2.5-flash", "gemini-2.5-pro", "gemini-2.5-flash-lite"]
    model_candidates = [model] + [m for m in fallback_order if m != model]

    encoded_key = urllib.parse.quote(GEMINI_API_KEY, safe="")
    ssl_context = ssl.create_default_context(cafile=certifi.where())
    last_error = None
    for candidate in model_candidates:
        endpoint = f"https://generativelanguage.googleapis.com/v1beta/models/{candidate}:generateContent?key={encoded_key}"
        payload = {
            "contents": [{"parts": [{"text": prompt}]}],
            "generationConfig": {
                "temperature": temperature,
                "maxOutputTokens": max_tokens,
            },
        }
        request = urllib.request.Request(
            endpoint,
            data=json.dumps(payload).encode("utf-8"),
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        try:
            with urllib.request.urlopen(request, timeout=60, context=ssl_context) as response:
                text = response.read().decode("utf-8")
                result = json.loads(text)
        except urllib.error.HTTPError as exc:
            error_text = exc.read().decode("utf-8", errors="replace")
            try:
                error_json = json.loads(error_text)
                message = error_json.get("error", {}).get("message", error_text)
            except Exception:
                message = error_text
            last_error = RuntimeError(f"Gemini API request failed for {candidate} (v1beta): {message}")
            if exc.code == 404 or "not found" in message.lower():
                continue
            raise last_error from exc
        except Exception as exc:
            last_error = RuntimeError(f"Gemini API request failed for {candidate} (v1beta): {exc}")
            continue

        if not result:
            last_error = RuntimeError("Gemini response was empty")
            continue

        # Modern Gemini generateContent response shape.
        candidates = result.get("candidates") or []
        if candidates:
            content = candidates[0].get("content", {})
            parts = content.get("parts", []) if isinstance(content, dict) else []
            texts = [p.get("text", "") for p in parts if isinstance(p, dict) and p.get("text")]
            if texts:
                return "".join(texts).strip()
            finish_reason = candidates[0].get("finishReason")
            if finish_reason:
                last_error = RuntimeError(f"Gemini returned no text for {candidate} (finishReason: {finish_reason})")
                continue

        prompt_feedback = result.get("promptFeedback", {})
        if prompt_feedback.get("blockReason"):
            last_error = RuntimeError(f"Gemini blocked the prompt for {candidate}: {prompt_feedback.get('blockReason')}")
            continue

        last_error = RuntimeError(f"Unable to parse Gemini response from {candidate} (v1beta)")

    if last_error is not None:
        raise last_error
    raise RuntimeError("Gemini API request failed: no usable response returned")


def _groq_chat(prompt: str, model: str = "qwen/qwen3-32b", temperature: float = 0.2, max_tokens: int = 600) -> str:
    if not GROQ_API_KEY:
        raise RuntimeError("Groq API key is not configured. Set GROQ_API_KEY in .env")

    endpoint = "https://api.groq.com/openai/v1/chat/completions"
    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": "You are a concise and factual financial crime investigation assistant."},
            {"role": "user", "content": prompt},
        ],
        "temperature": temperature,
        "max_tokens": max_tokens,
    }
    ssl_context = ssl.create_default_context(cafile=certifi.where())
    request = urllib.request.Request(
        endpoint,
        data=json.dumps(payload).encode("utf-8"),
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {GROQ_API_KEY}",
            "User-Agent": "Mozilla/5.0 (compatible; StreamlitApp/1.0)",
            "Accept": "application/json",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=60, context=ssl_context) as response:
            text = response.read().decode("utf-8")
            result = json.loads(text)
    except urllib.error.HTTPError as exc:
        error_text = exc.read().decode("utf-8", errors="replace")
        try:
            error_json = json.loads(error_text)
            message = error_json.get("error", {}).get("message", error_text)
        except Exception:
            message = error_text
        raise RuntimeError(f"Groq API request failed for {model}: {message}") from exc
    except Exception as exc:
        raise RuntimeError(f"Groq API request failed for {model}: {exc}") from exc

    choices = result.get("choices") or []
    if choices:
        message = choices[0].get("message", {})
        content = message.get("content", "")
        if content:
            return str(content).strip()

    raise RuntimeError(f"Unable to parse Groq response from {model}")


def _generate_groq_investigator_summary(target_context: dict, baseline_text: str, model: str = "qwen/qwen3-32b") -> str:
    prompt = (
        "Rewrite the following investigator summary in a clear, concise, factual way for an AML investigator. "
        "Keep it short, professional, and actionable. Do not invent facts.\n\n"
        f"Baseline summary:\n{baseline_text}\n\n"
        "Facts:\n"
        f"- Target customer: {target_context.get('target_customer_id', '')}\n"
        f"- Selected network: {target_context.get('selected_network', '')}\n"
        f"- Lookback window: {target_context.get('window_label', '')}\n"
        f"- Network score: {float(target_context.get('network_score', 0.0)):.2f}\n"
        f"- Node score: {float(target_context.get('node_score', 0.0)):.2f}\n"
        f"- Risk band: {target_context.get('risk_band', '')}\n"
        f"- Transactions: {int(target_context.get('txn_count', 0))}\n"
        f"- Total USD: {float(target_context.get('total_value', 0.0)):.2f}\n"
        f"- Outgoing count: {int(target_context.get('outgoing_count', 0))}\n"
        f"- Incoming count: {int(target_context.get('incoming_count', 0))}\n"
        f"- Top themes: {target_context.get('top_theme_text', '')}\n"
        f"- Flags: {', '.join(target_context.get('flag_text', [])) if target_context.get('flag_text') else 'none'}\n"
        f"- Top drivers: {target_context.get('top_driver_text', '')}\n"
    )
    return _groq_chat(prompt, model=model)


def _generate_groq_node_risk_summary(node_context: dict, baseline_text: str, model: str = "qwen/qwen3-32b") -> str:
    prompt = (
        "Rewrite the following node risk summary in a concise, investigator-ready way. "
        "Use only the facts provided and keep the language clear and factual.\n\n"
        f"Baseline summary:\n{baseline_text}\n\n"
        "Node facts:\n"
        f"- Selected node: {node_context.get('selected_node', '')}\n"
        f"- Network: {node_context.get('selected_network', '')}\n"
        f"- Node risk score: {float(node_context.get('node_score', 0.0)):.2f}\n"
        f"- Network score: {float(node_context.get('network_score', 0.0)):.2f}\n"
        f"- Risk band: {node_context.get('risk_band', '')}\n"
        f"- Transactions: {int(node_context.get('txn_count', 0))}\n"
        f"- Total USD: {float(node_context.get('total_value', 0.0)):.2f}\n"
        f"- Outgoing count: {int(node_context.get('outgoing_count', 0))}\n"
        f"- Incoming count: {int(node_context.get('incoming_count', 0))}\n"
        f"- Top drivers: {node_context.get('top_driver_text', '')}\n"
        f"- Flags: {', '.join(node_context.get('flag_text', [])) if node_context.get('flag_text') else 'none'}\n"
        f"- Themes: {node_context.get('top_theme_text', '')}\n"
        f"- Counterparties: {node_context.get('top_counterparty_text', '')}\n"
    )
    return _groq_chat(prompt, model=model)


def _generate_groq_next_actions(node_context: dict, fallback_actions: List[str], model: str = "qwen/qwen3-32b") -> List[str]:
    prompt = (
        "You are an AML investigator assistant. Review the facts below and generate 5 concise next investigative actions. "
        "Return only the actions as a numbered list.\n\n"
        "Node facts:\n"
        f"- Selected node: {node_context.get('selected_node', '')}\n"
        f"- Network: {node_context.get('selected_network', '')}\n"
        f"- Node risk score: {float(node_context.get('node_score', 0.0)):.2f}\n"
        f"- Network score: {float(node_context.get('network_score', 0.0)):.2f}\n"
        f"- Transaction count: {int(node_context.get('txn_count', 0))}\n"
        f"- Total USD: {float(node_context.get('total_value', 0.0)):.2f}\n"
        f"- Top drivers: {node_context.get('top_driver_text', '')}\n"
        f"- Flags: {', '.join(node_context.get('flag_text', [])) if node_context.get('flag_text') else 'none'}\n"
        f"- Themes: {node_context.get('top_theme_text', '')}\n"
        f"- Counterparties: {node_context.get('top_counterparty_text', '')}\n\n"
        "Baseline investigator actions:\n"
        + "\n".join([f"- {a}" for a in fallback_actions])
    )
    result = _groq_chat(prompt, model=model, max_tokens=1200)
    return _format_action_lines(result)


def build_network_recommendation_prompt(n_hops: int, lookback_days: int) -> str:
    return (
        "You are a Financial Crime Network Analyst. Analyse the data shared, your task is to find out networks "
        f"up to {int(n_hops)} hops and lookback {int(lookback_days)} days using SAR or exit or Sanction or pep or "
        "DRA alert flag where you see financial crime exposure and recommend strict action along with rationale "
        "or when there is no significant risk conclude where we need to do enhanced due diligence and rationale. "
        "Provide connected customers information wherever action needs to be taken with rationale."
        "Keep the response concise, under 1000 words, and end with a complete sentence"
    )

DRA_ALERT_SCORE_THRESHOLD = 970.0


def _build_network_flag_context(analyzer, outputs: dict, target_customer_id: str) -> str:
    """Builds a factual, model-ready summary of networks and SAR/Sanctions/PEP/Exit/DRA-alert flagged
    entities discovered within the configured hop radius and lookback window, for use as grounding
    context in AI recommendation prompts."""
    node_rankings = outputs.get("node_rankings", pd.DataFrame())
    network_summary = outputs.get("network_summary", pd.DataFrame())

    if node_rankings.empty or network_summary.empty:
        return (
            f"Target customer: {target_customer_id}\n"
            f"Lookback window: {getattr(analyzer, 'window_label', '')}\n"
            "No networks or nodes were found for the selected target within the configured lookback and hop settings."
        )

    profiles = getattr(analyzer, "node_profiles", pd.DataFrame())
    dra_map = {}
    if isinstance(profiles, pd.DataFrame) and not profiles.empty and "dra_score" in profiles.columns:
        dra_map = dict(zip(profiles["customer_id"].astype(str), profiles["dra_score"]))

    network_lines = []
    for r in network_summary.itertuples(index=False):
        network_lines.append(
            f"- Network {r.network_id}: {r.nodes} nodes, {r.edges} edges, {r.txn_count} txns, "
            f"total ${r.total_amount_usd:,.0f}, cross-border ratio {r.cross_border_ratio:.2f}, "
            f"flagged nodes ({r.flagged_nodes_counts}), network risk score {r.network_risk_score:.1f}"
        )

    flagged_rows = []
    for r in node_rankings.itertuples(index=False):
        dra_score = float(dra_map.get(str(r.customer_id), 0.0) or 0.0)
        dra_alert = dra_score >= DRA_ALERT_SCORE_THRESHOLD
        if bool(r.sanctions_flag) or bool(r.pep_flag) or bool(r.sar_flag) or bool(r.exited_flag) or dra_alert:
            flags = []
            if r.sanctions_flag:
                flags.append("Sanctions")
            if r.sar_flag:
                flags.append("SAR")
            if r.pep_flag:
                flags.append("PEP")
            if r.exited_flag:
                flags.append("Exited")
            if dra_alert:
                flags.append(f"DRA score ({dra_score:.1f})")
            flagged_rows.append(
                f"- Network {r.network_id} | Customer {r.customer_id} ({r.customer_name}) | "
                f"Node risk score {r.final_node_risk_score:.1f} | Flags: {', '.join(flags)}"
            )

    context = (
        f"Target customer: {target_customer_id}\n"
        f"Lookback window: {getattr(analyzer, 'window_label', '')}\n"
        f"Hop radius analyzed: up to {analyzer.config.n_hops} hops\n\n"
        "Network-level summary:\n"
        + ("\n".join(network_lines) if network_lines else "No networks discovered.")
        + "\n\n"
        "Flagged entities (SAR / Sanctions / PEP / Exited / DRA Alert) within scope:\n"
        + (
            "\n".join(flagged_rows)
            if flagged_rows
            else "No SAR, Sanctions, PEP, Exited, or high-DRA-score entities found within the analyzed hop radius."
        )
    )
    return context


def _ollama_chat(prompt: str, model: str = "llama3.1:8b", temperature: float = 0.2) -> str:
    resp = ollama.chat(
        model=model,
        messages=[{"role": "user", "content": prompt}],
        options={"temperature": temperature},
    )
    content = str(resp.get("message", {}).get("content", "") or "").strip()
    if not content:
        raise RuntimeError("Ollama returned an empty response")
    return content


def _build_deterministic_network_recommendation(context_text: str, n_hops: int, lookback_days: int) -> str:
    """Rule-based fallback recommendation used when no AI provider is available or the call fails."""
    has_flags = "No SAR, Sanctions, PEP, Exited, or high-DRA-score entities" not in context_text
    if has_flags:
        return (
            "**Deterministic assessment (AI unavailable):** One or more SAR, Sanctions, PEP, Exit, or "
            f"high-DRA-score entities were identified within {int(n_hops)} hops of the target customer in the "
            f"last {int(lookback_days)} days. See the flagged entities and their connected network/customer IDs "
            "listed above.\n\n"
            "**Recommended action:** Escalate for strict review — freeze/hold outbound activity pending review, "
            "re-screen the target and the connected customers listed above against current sanctions/PEP/adverse-media "
            "lists, and file or update a SAR if not already in progress.\n\n"
            f"**Rationale:** Proximity (within {int(n_hops)} hops) to flagged entities materially raises facilitation "
            "and layering risk; the specific flagged entities, their connected network, and hop distance are listed "
            "above and should anchor the investigative narrative."
        )
    return (
        "**Deterministic assessment (AI unavailable):** No SAR, Sanctions, PEP, Exit, or high-DRA-score entities "
        f"were found within {int(n_hops)} hops of the target customer in the last {int(lookback_days)} days.\n\n"
        "**Recommended action:** No strict action warranted at this time; proceed with Enhanced Due Diligence (EDD) "
        "— verify source of funds/wealth, confirm beneficial ownership, and monitor for behavioural drift over the "
        "next review cycle.\n\n"
        f"**Rationale:** Absence of direct or {int(n_hops)}-hop-proximate flags lowers immediate escalation priority, "
        "but EDD maintains a documented risk-based control given the customer was still selected for network review."
    )


def generate_ai_network_recommendation(
    analyzer,
    outputs: dict,
    target_customer_id: str,
    ai_provider: str,
    ai_model: str,
    ollama_model: str,
) -> tuple:
    """Runs the Financial Crime Network Analyst prompt against the selected AI provider, grounded in the
    actual network data, using the hop radius and lookback window configured in the sidebar for this run.
    Returns (recommendation_text, status_message)."""
    n_hops = int(getattr(analyzer.config, "n_hops", 2))
    lookback_days = int(getattr(analyzer.config, "lookback_days", 180))

    context_text = _build_network_flag_context(analyzer, outputs, target_customer_id)
    analyst_prompt = build_network_recommendation_prompt(n_hops=n_hops, lookback_days=lookback_days)
    full_prompt = f"{analyst_prompt}\n\n=== DATA ===\n{context_text}"

    try:
        if ai_provider == "OpenAI":
            if not OPENAI_API_KEY:
                raise RuntimeError("OPENAI_API_KEY is not configured")
            text = _openai_chat(full_prompt, model=ai_model or "gpt-5.5", max_tokens=1200)
            status = f"AI Model Recommendation generated using OpenAI model '{ai_model}'."
        elif ai_provider == "Gemini":
            if not GEMINI_API_KEY:
                raise RuntimeError("GEMINI_API_KEY is not configured")
            text = _gemini_chat(full_prompt, model=ai_model or "gemini-2.5-flash", max_tokens=1200)
            status = f"AI Model Recommendation generated using Gemini model '{ai_model}'."
        elif ai_provider == "Groq":
            if not GROQ_API_KEY:
                raise RuntimeError("GROQ_API_KEY is not configured")
            text = _groq_chat(full_prompt, model=ai_model or "qwen/qwen3-32b", max_tokens=1200)
            status = f"AI Model Recommendation generated using Groq model '{ai_model}'."
        elif ai_provider == "Ollama (Local)":
            if not HAS_OLLAMA:
                raise RuntimeError("ollama package is not installed")
            text = _ollama_chat(full_prompt, model=ollama_model.strip() or "llama3.1:8b")
            status = f"AI Model Recommendation generated using Ollama model '{(ollama_model.strip() or 'llama3.1:8b')}'."
        else:
            text = _build_deterministic_network_recommendation(context_text, n_hops=n_hops, lookback_days=lookback_days)
            status = "AI provider set to Deterministic; showing rule-based network recommendation."
        # return text.strip(), status
        return _trim_incomplete_tail(text), status
    
    except Exception as ex:
        fallback_text = _build_deterministic_network_recommendation(context_text, n_hops=n_hops, lookback_days=lookback_days)
        status = f"{ai_provider} unavailable for AI Model Recommendation ({ex}); showing deterministic recommendation."
        return fallback_text, status


def _generate_openai_investigator_summary(target_context: dict, baseline_text: str, model: str = "gpt-5.5") -> str:
    prompt = (
        "Rewrite the following investigator summary in a clear, concise, factual way for an AML investigator. "
        "Keep it short, professional, and actionable. Do not invent facts.\n\n"
        f"Baseline summary:\n{baseline_text}\n\n"
        "Facts:\n"
        f"- Target customer: {target_context.get('target_customer_id', '')}\n"
        f"- Selected network: {target_context.get('selected_network', '')}\n"
        f"- Lookback window: {target_context.get('window_label', '')}\n"
        f"- Network score: {float(target_context.get('network_score', 0.0)):.2f}\n"
        f"- Node score: {float(target_context.get('node_score', 0.0)):.2f}\n"
        f"- Risk band: {target_context.get('risk_band', '')}\n"
        f"- Transactions: {int(target_context.get('txn_count', 0))}\n"
        f"- Total USD: {float(target_context.get('total_value', 0.0)):.2f}\n"
        f"- Outgoing count: {int(target_context.get('outgoing_count', 0))}\n"
        f"- Incoming count: {int(target_context.get('incoming_count', 0))}\n"
        f"- Top themes: {target_context.get('top_theme_text', '')}\n"
        f"- Flags: {', '.join(target_context.get('flag_text', [])) if target_context.get('flag_text') else 'none'}\n"
        f"- Top drivers: {target_context.get('top_driver_text', '')}\n"
    )
    return _openai_chat(prompt, model=model)


def _generate_openai_node_risk_summary(node_context: dict, baseline_text: str, model: str = "gpt-5.5") -> str:
    prompt = (
        "Rewrite the following node risk summary in a concise, investigator-ready way. "
        "Use only the facts provided and keep the language clear and factual.\n\n"
        f"Baseline summary:\n{baseline_text}\n\n"
        "Node facts:\n"
        f"- Selected node: {node_context.get('selected_node', '')}\n"
        f"- Network: {node_context.get('selected_network', '')}\n"
        f"- Node risk score: {float(node_context.get('node_score', 0.0)):.2f}\n"
        f"- Network score: {float(node_context.get('network_score', 0.0)):.2f}\n"
        f"- Risk band: {node_context.get('risk_band', '')}\n"
        f"- Transactions: {int(node_context.get('txn_count', 0))}\n"
        f"- Total USD: {float(node_context.get('total_value', 0.0)):.2f}\n"
        f"- Outgoing count: {int(node_context.get('outgoing_count', 0))}\n"
        f"- Incoming count: {int(node_context.get('incoming_count', 0))}\n"
        f"- Top drivers: {node_context.get('top_driver_text', '')}\n"
        f"- Flags: {', '.join(node_context.get('flag_text', [])) if node_context.get('flag_text') else 'none'}\n"
        f"- Themes: {node_context.get('top_theme_text', '')}\n"
        f"- Counterparties: {node_context.get('top_counterparty_text', '')}\n"
    )
    return _openai_chat(prompt, model=model)


def _generate_openai_next_actions(node_context: dict, fallback_actions: List[str], model: str = "gpt-5.5") -> List[str]:
    prompt = (
        "You are an AML investigator assistant. Review the facts below and generate 5 concise next investigative actions. "
        "Return only the actions as a numbered list.\n\n"
        "Node facts:\n"
        f"- Selected node: {node_context.get('selected_node', '')}\n"
        f"- Network: {node_context.get('selected_network', '')}\n"
        f"- Node risk score: {float(node_context.get('node_score', 0.0)):.2f}\n"
        f"- Network score: {float(node_context.get('network_score', 0.0)):.2f}\n"
        f"- Transaction count: {int(node_context.get('txn_count', 0))}\n"
        f"- Total USD: {float(node_context.get('total_value', 0.0)):.2f}\n"
        f"- Top drivers: {node_context.get('top_driver_text', '')}\n"
        f"- Flags: {', '.join(node_context.get('flag_text', [])) if node_context.get('flag_text') else 'none'}\n"
        f"- Themes: {node_context.get('top_theme_text', '')}\n"
        f"- Counterparties: {node_context.get('top_counterparty_text', '')}\n\n"
        "Baseline investigator actions:\n"
        + "\n".join([f"- {a}" for a in fallback_actions])
    )
    result = _openai_chat(prompt, model=model, max_tokens=1200)
    return _format_action_lines(result)


def _generate_gemini_investigator_summary(target_context: dict, baseline_text: str, model: str = "gemini-2.5-flash") -> str:
    prompt = (
        "Rewrite the following investigator summary in a clear, concise, factual way for an AML investigator. "
        "Keep it short, professional, and actionable. Do not invent facts.\n\n"
        f"Baseline summary:\n{baseline_text}\n\n"
        "Facts:\n"
        f"- Target customer: {target_context.get('target_customer_id', '')}\n"
        f"- Selected network: {target_context.get('selected_network', '')}\n"
        f"- Lookback window: {target_context.get('window_label', '')}\n"
        f"- Network score: {float(target_context.get('network_score', 0.0)):.2f}\n"
        f"- Node score: {float(target_context.get('node_score', 0.0)):.2f}\n"
        f"- Risk band: {target_context.get('risk_band', '')}\n"
        f"- Transactions: {int(target_context.get('txn_count', 0))}\n"
        f"- Total USD: {float(target_context.get('total_value', 0.0)):.2f}\n"
        f"- Outgoing count: {int(target_context.get('outgoing_count', 0))}\n"
        f"- Incoming count: {int(target_context.get('incoming_count', 0))}\n"
        f"- Top themes: {target_context.get('top_theme_text', '')}\n"
        f"- Flags: {', '.join(target_context.get('flag_text', [])) if target_context.get('flag_text') else 'none'}\n"
        f"- Top drivers: {target_context.get('top_driver_text', '')}\n"
    )
    return _gemini_chat(prompt, model=model)


def _generate_gemini_node_risk_summary(node_context: dict, baseline_text: str, model: str = "gemini-2.5-flash") -> str:
    prompt = (
        "Rewrite the following node risk summary in a concise, investigator-ready way. "
        "Use only the facts provided and keep the language clear and factual.\n\n"
        f"Baseline summary:\n{baseline_text}\n\n"
        "Node facts:\n"
        f"- Selected node: {node_context.get('selected_node', '')}\n"
        f"- Network: {node_context.get('selected_network', '')}\n"
        f"- Node risk score: {float(node_context.get('node_score', 0.0)):.2f}\n"
        f"- Network score: {float(node_context.get('network_score', 0.0)):.2f}\n"
        f"- Risk band: {node_context.get('risk_band', '')}\n"
        f"- Transactions: {int(node_context.get('txn_count', 0))}\n"
        f"- Total USD: {float(node_context.get('total_value', 0.0)):.2f}\n"
        f"- Outgoing count: {int(node_context.get('outgoing_count', 0))}\n"
        f"- Incoming count: {int(node_context.get('incoming_count', 0)):.2f}\n"
        f"- Top drivers: {node_context.get('top_driver_text', '')}\n"
        f"- Flags: {', '.join(node_context.get('flag_text', [])) if node_context.get('flag_text') else 'none'}\n"
        f"- Themes: {node_context.get('top_theme_text', '')}\n"
        f"- Counterparties: {node_context.get('top_counterparty_text', '')}\n"
    )
    return _gemini_chat(prompt, model=model)


def _generate_gemini_next_actions(node_context: dict, fallback_actions: List[str], model: str = "gemini-2.5-flash") -> List[str]:
    prompt = (
        "You are an AML investigator assistant. Review the facts below and generate 5 concise next investigative actions. "
        "Return only the actions as a numbered list.\n\n"
        "Node facts:\n"
        f"- Selected node: {node_context.get('selected_node', '')}\n"
        f"- Network: {node_context.get('selected_network', '')}\n"
        f"- Node risk score: {float(node_context.get('node_score', 0.0)):.2f}\n"
        f"- Network score: {float(node_context.get('network_score', 0.0)):.2f}\n"
        f"- Transaction count: {int(node_context.get('txn_count', 0))}\n"
        f"- Total USD: {float(node_context.get('total_value', 0.0)):.2f}\n"
        f"- Top drivers: {node_context.get('top_driver_text', '')}\n"
        f"- Flags: {', '.join(node_context.get('flag_text', [])) if node_context.get('flag_text') else 'none'}\n"
        f"- Themes: {node_context.get('top_theme_text', '')}\n"
        f"- Counterparties: {node_context.get('top_counterparty_text', '')}\n\n"
        "Baseline investigator actions:\n"
        + "\n".join([f"- {a}" for a in fallback_actions])
    )
    result = _gemini_chat(prompt, model=model, max_tokens=1200)
    return _format_action_lines(result)


def load_data() -> pd.DataFrame:
    base_dir = Path(__file__).resolve().parent
    candidates = []
    for name in (
        "synthetic_transactions_complex_INTERNAL.xlsx",
        "synthetic_transactions_complex_INTERNAL.csv",
        "synthetic_transactions_complex_INTERNAL.xls",
        "synthetic_transactions_complex_INTERNAL.xlsm",
    ):
        candidates.append(base_dir / name)
        candidates.append(Path.cwd() / name)

    seen_paths = set()
    last_error = None
    for path in candidates:
        if not path.exists() or path in seen_paths:
            continue
        seen_paths.add(path)

        try:
            suffix = path.suffix.lower()
            if suffix in {".xlsx", ".xls", ".xlsm", ".xlsb"}:
                try:
                    import openpyxl  # noqa: F401
                except ImportError as exc:
                    raise ImportError("openpyxl is required to read Excel files") from exc
                return pd.read_excel(path, engine="openpyxl")

            with open(path, "rb") as handle:
                header = handle.read(4)
            if header.startswith(b"PK\x03\x04"):
                try:
                    import openpyxl  # noqa: F401
                except ImportError as exc:
                    raise ImportError("openpyxl is required to read Excel files") from exc
                return pd.read_excel(path, engine="openpyxl")

            with open(path, "r", encoding="utf-8-sig", errors="replace", newline="") as handle:
                reader = csv.reader(handle, quotechar='"', escapechar="\\")
                rows = [row for row in reader if any(cell.strip() for cell in row)]
                if not rows:
                    raise ValueError("CSV file is empty or could not be parsed")
                header = [col.strip() for col in rows[0]]
                rows_data = np.array(rows[1:], dtype=object)
                return pd.DataFrame(rows_data, columns=header)
        except Exception as exc:  # pragma: no cover - defensive fallback
            last_error = exc

    if last_error is not None:
        raise RuntimeError(f"Unable to read dataset: {last_error}") from last_error

    raise FileNotFoundError("No transaction dataset file was found in the project directory")


def parse_mode(mode: str) -> str:
    return "value" if mode == "value-weighted" else "count"


def estimate_unique_nodes_within_hops(
    df: pd.DataFrame,
    target_customer_id: str,
    use_all_dates: bool,
    lookback_days: int,
    n_hops: int,
) -> int:
    if df is None or df.empty:
        return 0

    working_df = df
    if not bool(use_all_dates) and "transaction_datetime" in df.columns:
        dt_series = pd.to_datetime(df["transaction_datetime"], errors="coerce", utc=True).dt.tz_localize(None)
        valid_mask = dt_series.notna()
        if valid_mask.any():
            max_dt = dt_series[valid_mask].max()
            cutoff_dt = max_dt - pd.Timedelta(days=int(lookback_days))
            working_df = df.loc[valid_mask & (dt_series >= cutoff_dt)].copy()
        else:
            working_df = df.iloc[0:0].copy()

    if working_df.empty:
        return 0

    graph = nx.Graph()
    origins = working_df["originator_id"].astype(str)
    beneficiaries = working_df["beneficiary_id"].astype(str)
    graph.add_edges_from(zip(origins, beneficiaries))

    target_id = str(target_customer_id)
    if target_id not in graph:
        return 0

    reachable_nodes = nx.single_source_shortest_path_length(graph, target_id, cutoff=int(n_hops))
    return int(len(reachable_nodes))


def generate_ollama_customer_brief(
    model_name: str,
    selected_network: str,
    selected_node: str,
    node_score: float,
    network_score: float,
    txn_count: int,
    total_value: float,
    avg_value: float,
    outgoing_count: int,
    incoming_count: int,
    top_driver_text: str,
    flag_text,
    theme_names,
):
    prompt = "\n".join([
        "You are a financial crime investigator assistant.",
        "Return plain text only in this exact structure:",
        "SUMMARY:",
        "<2-3 short paragraphs>",
        "ACTIONS:",
        "- <action 1>",
        "- <action 2>",
        "- <action 3>",
        "- <action 4>",
        "- <action 5>",
        "",
        f"Network: {selected_network}",
        f"Customer: {selected_node}",
        f"Node score: {node_score:.2f}",
        f"Network score: {network_score:.2f}",
        f"Transactions: {txn_count}",
        f"Total USD: {total_value:,.2f}",
        f"Average USD: {avg_value:,.2f}",
        f"Outgoing count: {outgoing_count}",
        f"Incoming count: {incoming_count}",
        f"Top score drivers: {top_driver_text}",
        f"Direct flags: {', '.join(flag_text) if flag_text else 'none'}",
        f"Top themes: {', '.join(theme_names[:4]) if theme_names else 'none'}",
        "",
        "Keep it concise, factual, and actionable for AML investigation.",
    ])

    resp = ollama.chat(
        model=model_name,
        messages=[
            {
                "role": "user",
                "content": prompt,
            }
        ],
        options={"temperature": 0.2},
    )

    content = str(resp.get("message", {}).get("content", "") or "").strip()
    if not content:
        return "", []

    summary_text = ""
    actions = []
    lines = [ln.rstrip() for ln in content.splitlines()]

    in_summary = False
    in_actions = False
    summary_lines = []

    for ln in lines:
        clean = ln.strip()
        upper = clean.upper()
        if upper.startswith("SUMMARY:"):
            in_summary = True
            in_actions = False
            remainder = clean[len("SUMMARY:"):].strip()
            if remainder:
                summary_lines.append(remainder)
            continue
        if upper.startswith("ACTIONS:"):
            in_summary = False
            in_actions = True
            remainder = clean[len("ACTIONS:"):].strip()
            if remainder:
                if remainder.startswith("-"):
                    actions.append(remainder.lstrip("- ").strip())
                else:
                    actions.append(remainder)
            continue

        if in_summary:
            summary_lines.append(ln)
        elif in_actions:
            if clean.startswith("-"):
                item = clean.lstrip("- ").strip()
                if item:
                    actions.append(item)
            elif clean:
                actions.append(clean)

    if summary_lines:
        summary_text = "\n".join(summary_lines).strip()
    else:
        summary_text = content

    dedup_actions = []
    seen = set()
    for action in actions:
        key = action.lower()
        if key and key not in seen:
            seen.add(key)
            dedup_actions.append(action)

    return summary_text, dedup_actions[:6]


def generate_ollama_investigator_summary(
    model_name: str,
    target_customer_id: str,
    customer_name: str,
    selected_network: str,
    window_label: str,
    network_count: int,
    network_score: float,
    node_score: float,
    risk_band: str,
    txn_count: int,
    total_value: float,
    outgoing_count: int,
    incoming_count: int,
    top_driver_text: str,
    flag_text,
    top_theme_text: str,
    top_counterparty_text: str,
    top_scenario_text: str,
):
    prompt = "\n".join([
        "You are the AI Summary Generator for a financial crime investigator.",
        "Summarize the transaction network analysis for the targeted customer only.",
        "Use only the facts provided. Do not invent missing facts or overstate suspicion.",
        "Return plain text only in this exact structure:",
        "EXECUTIVE SUMMARY:",
        "<3-5 concise sentences>",
        "KEY FINDINGS:",
        "- <finding 1 with metric or evidence>",
        "- <finding 2 with metric or evidence>",
        "- <finding 3 with metric or evidence>",
        "EVIDENCE BASIS:",
        "<one concise sentence>",
        "LIMITATIONS:",
        "<one concise sentence>",
        "RECOMMENDED NEXT ACTIONS:",
        "- <action 1>",
        "- <action 2>",
        "- <action 3>",
        "",
        f"Lookback window: {window_label}",
        f"Target customer ID: {target_customer_id}",
        f"Target customer name: {customer_name or 'unknown'}",
        f"Selected network: {selected_network}",
        f"Networks discovered: {network_count}",
        f"Network risk score: {network_score:.2f}",
        f"Target node risk score: {node_score:.2f}",
        f"Target risk band: {risk_band}",
        f"Linked transactions for target in selected network: {txn_count}",
        f"Linked transaction value USD: {total_value:,.2f}",
        f"Outgoing count: {outgoing_count}",
        f"Incoming count: {incoming_count}",
        f"Top score drivers: {top_driver_text}",
        f"Direct flags: {', '.join(flag_text) if flag_text else 'none'}",
        f"Top network themes: {top_theme_text or 'none'}",
        f"Frequent counterparties: {top_counterparty_text or 'none'}",
        f"Scenario tags: {top_scenario_text or 'none'}",
    ])

    resp = ollama.chat(
        model=model_name,
        messages=[{"role": "user", "content": prompt}],
        options={"temperature": 0.2},
    )
    content = str(resp.get("message", {}).get("content", "") or "").strip()
    if not content:
        return "", []

    actions = []
    in_actions = False
    for ln in content.splitlines():
        clean = ln.strip()
        if clean.upper().startswith("RECOMMENDED NEXT ACTIONS"):
            in_actions = True
            continue
        if in_actions and clean.startswith("-"):
            item = clean.lstrip("- ").strip()
            if item:
                actions.append(item)

    return content, actions[:6]


def get_ollama_runtime_status(model_name: str):
    if not HAS_OLLAMA:
        return False, False, "Ollama Python package is not installed. Run: pip install ollama"

    desired = str(model_name or "").strip()
    try:
        resp = ollama.list()
        models_data = []
        if isinstance(resp, dict):
            models_data = resp.get("models", []) or []
        elif hasattr(resp, "models"):
            models_data = getattr(resp, "models", []) or []

        names = []
        for m in models_data:
            if isinstance(m, dict):
                name = str(m.get("model") or m.get("name") or "").strip()
            else:
                name = str(getattr(m, "model", "") or getattr(m, "name", "")).strip()
            if name:
                names.append(name)

        is_available = any(
            n == desired or n.startswith(f"{desired}:") or desired.startswith(f"{n}:")
            for n in names
        ) if desired else False

        if is_available:
            return True, True, f"Connected. Model available: {desired}"

        if names:
            shown = ", ".join(names[:4])
            suffix = "..." if len(names) > 4 else ""
            return True, False, f"Connected. Model not found: {desired}. Local models: {shown}{suffix}"

        return True, False, "Connected, but no local models found. Pull one with: ollama pull llama3.1:8b"
    except Exception as ex:
        return False, False, f"Ollama service is not reachable: {ex}"


NODE_SCORE_MAX = 58.5


def node_risk_band_from_internal(score: float):
    very_high_cutoff = 0.8 * NODE_SCORE_MAX
    high_cutoff = 0.6 * NODE_SCORE_MAX
    medium_cutoff = 0.4 * NODE_SCORE_MAX

    if score >= very_high_cutoff:
        return "🔴 Very High", "#b30000"
    if score >= high_cutoff:
        return "🟠 High", "#c95f02"
    if score >= medium_cutoff:
        return "🟡 Medium", "#856404"
    return "🟢 Low", "#2e7d32"


def format_customer_label(customer_id: str, customer_name: str) -> str:
    name = str(customer_name or "").strip()
    return f"{customer_id} - {name}" if name else str(customer_id)


def graph_color(sanctions: bool, pep: bool, sar: bool, exited: bool, is_target: bool) -> str:
    if is_target:
        return "#1f4aff"
    if sanctions:
        return "#d73027"
    if sar:
        return "#7a0177"
    if pep:
        return "#f39c12"
    if exited:
        return "#00a6a6"
    return "#4d4d4d"


def edge_risk_color(risk_score: float) -> str:
    """Colors an edge/arrow by the same Low/Medium/High/Very High risk-band scale used
    throughout the app's risk summaries, so risky transaction paths and sub-networks are
    visually consistent with the written risk assessment."""
    _, color = node_risk_band_from_internal(float(risk_score or 0.0))
    return color


def _hex_to_rgba(color_hex: str, alpha: float) -> str:
    color_hex = str(color_hex or "").strip().lstrip("#")
    if len(color_hex) != 6:
        return f"rgba(148, 163, 184, {alpha})"
    r = int(color_hex[0:2], 16)
    g = int(color_hex[2:4], 16)
    b = int(color_hex[4:6], 16)
    return f"rgba({r}, {g}, {b}, {alpha})"


def _cluster_highlight_color(cluster_id: int) -> str:
    palette = [
        "#60a5fa", "#34d399", "#f59e0b", "#f472b6", "#a78bfa",
        "#22d3ee", "#fb7185", "#86efac", "#fca5a5", "#c4b5fd",
    ]
    if int(cluster_id) <= 0:
        return "#1f4aff"
    return palette[(int(cluster_id) - 1) % len(palette)]


def build_network_layout(node_ids, edge_df: pd.DataFrame, scale: float):
    graph = nx.Graph()
    graph.add_nodes_from([str(n) for n in node_ids])
    for r in edge_df.itertuples(index=False):
        graph.add_edge(str(r.originator_id), str(r.beneficiary_id), weight=max(1.0, float(r.txn_count)))

    if graph.number_of_nodes() == 0:
        return {}
    if graph.number_of_nodes() == 1:
        only_node = next(iter(graph.nodes()))
        return {only_node: (0.0, 0.0)}

    positions = nx.spring_layout(
        graph,
        weight="weight",
        seed=42,
        k=1.0 / max(1.0, graph.number_of_nodes() ** 0.5),
        iterations=80,
    )
    return {str(n): (float(x) * scale, float(y) * scale) for n, (x, y) in positions.items()}


def build_graph_objects(node_df: pd.DataFrame, edge_df: pd.DataFrame, target_customer_id: str):
    nodes = []
    node_ids = node_df["customer_id"].astype(str).tolist()
    positions = build_network_layout(node_ids, edge_df, scale=360.0)

    for r in node_df.itertuples(index=False):
        cid = str(r.customer_id)
        px, py = positions.get(cid, (0.0, 0.0))
        dra_score = float(getattr(r, "dra_score", 0.0) or 0.0)
        title = (
            f"ID: {r.customer_id}\n"
            f"Name: {r.customer_name}\n"
            f"Country: {r.country}\n"
            f"Entity: {r.entity_type}\n"
            f"Risk Score: {r.final_node_risk_score}\n"
            f"DRA Score: {dra_score:.2f}\n"
            f"Sanctions: {bool(r.sanctions_flag)} | PEP: {bool(r.pep_flag)} | SAR: {bool(r.sar_flag)} | Exited: {bool(r.exited_flag)}"
        )
        nodes.append(
            Node(
                id=str(r.customer_id),
                label="",
                title=title,
                size=10 + min(25, float(r.final_node_risk_score) / 4.0),
                x=px,
                y=py,
                color=graph_color(
                    bool(r.sanctions_flag),
                    bool(r.pep_flag),
                    bool(r.sar_flag),
                    bool(r.exited_flag),
                    cid == str(target_customer_id),
                ),
            )
        )

    node_score_map = {
        str(r.customer_id): float(r.final_node_risk_score) for r in node_df.itertuples(index=False)
    }

    edges = []
    for r in edge_df.itertuples(index=False):
        src_id = str(r.originator_id)
        dst_id = str(r.beneficiary_id)
        edge_score = max(node_score_map.get(src_id, 0.0), node_score_map.get(dst_id, 0.0))
        edges.append(
            Edge(
                source=src_id,
                target=dst_id,
                width=1 + min(8, float(r.txn_count) / 2.0),
                color=edge_risk_color(edge_score),
                arrows="to",
            )
        )

    return nodes, edges


def _dot_escape(value: str) -> str:
    return str(value).replace("\\", "\\\\").replace('"', '\\"')


def build_graphviz_dot(node_df: pd.DataFrame, edge_df: pd.DataFrame, target_customer_id: str) -> str:
    lines = [
        "digraph TransactionNetwork {",
        "rankdir=LR;",
        'graph [bgcolor="white", overlap=false, splines=true, pad=0.2, nodesep=0.35, ranksep=0.6];',
        'node [shape=ellipse, style="filled", fontname="Helvetica", fontsize=10, color="#333333"];',
        'edge [fontname="Helvetica", fontsize=9, color="#666666", arrowsize=0.8];',
    ]

    for r in node_df.itertuples(index=False):
        node_id = _dot_escape(str(r.customer_id))
        color = graph_color(
            bool(r.sanctions_flag),
            bool(r.pep_flag),
            bool(r.sar_flag),
            bool(r.exited_flag),
            str(r.customer_id) == str(target_customer_id),
        )
        risk = float(r.final_node_risk_score)
        width = 0.5 + min(1.1, risk / 100.0)
        dra_score = float(getattr(r, "dra_score", 0.0) or 0.0)
        label = _dot_escape(f"{r.customer_id}\nDRA {dra_score:.2f}")
        lines.append(
            f'"{node_id}" [label="{label}", fillcolor="{color}", width={width:.2f}, height={width:.2f}, fixedsize=true];'
        )

    node_score_map = {
        str(r.customer_id): float(r.final_node_risk_score) for r in node_df.itertuples(index=False)
    }

    for r in edge_df.itertuples(index=False):
        src = _dot_escape(str(r.originator_id))
        dst = _dot_escape(str(r.beneficiary_id))
        src_id = str(r.originator_id)
        dst_id = str(r.beneficiary_id)
        edge_score = max(node_score_map.get(src_id, 0.0), node_score_map.get(dst_id, 0.0))
        edge_color = edge_risk_color(edge_score)
        penwidth = 1.0 + min(4.0, float(r.txn_count) / 3.0)
        lines.append(f'"{src}" -> "{dst}" [color="{edge_color}", penwidth={penwidth:.2f}];')

    lines.append("}")
    return "\n".join(lines)


def build_plotly_network_figure(
    node_df: pd.DataFrame,
    edge_df: pd.DataFrame,
    target_customer_id: str,
    cluster_map: dict | None = None,
    active_cluster_id: int | None = None,
):
    ordered_rows = list(node_df.itertuples(index=False))
    if not ordered_rows:
        return None

    target_id = str(target_customer_id)
    ordered_ids = [str(r.customer_id) for r in ordered_rows]
    if target_id not in ordered_ids:
        target_id = ordered_ids[0]
    pos = build_network_layout(ordered_ids, edge_df, scale=4.5)

    id_to_row = {str(r.customer_id): r for r in ordered_rows}

    node_x, node_y, node_text, node_sizes, node_colors, node_ids = [], [], [], [], [], []
    for n in ordered_ids:
        r = id_to_row[n]
        x, y = pos[n]
        cluster_id = int(cluster_map.get(str(n), -1)) if isinstance(cluster_map, dict) and cluster_map else -1
        is_active_node = (
            active_cluster_id is None
            or str(n) == str(target_customer_id)
            or cluster_id == int(active_cluster_id)
        )
        node_x.append(x)
        node_y.append(y)
        node_text.append("<br>".join([
            f"<b>{n}</b>",
            f"Name: {r.customer_name}",
            f"Country: {r.country} | Entity: {r.entity_type}",
            f"Risk Score: <b>{float(r.final_node_risk_score):.2f}</b>",
            f"DRA Score: <b>{float(getattr(r, 'dra_score', 0.0) or 0.0):.2f}</b>",
            f"Sanctions: {bool(r.sanctions_flag)} | PEP: {bool(r.pep_flag)} | SAR: {bool(r.sar_flag)} | Exited: {bool(r.exited_flag)}",
        ]))
        node_sizes.append(14 + min(30, float(r.final_node_risk_score) / 3.0))
        if is_active_node:
            node_colors.append(graph_color(
                bool(r.sanctions_flag), bool(r.pep_flag), bool(r.sar_flag),
                bool(r.exited_flag), str(n) == str(target_customer_id),
            ))
        else:
            node_colors.append("#b8bec8")
        node_ids.append(n)

    node_trace = go.Scatter(
        x=node_x, y=node_y,
        mode="markers",
        hoverinfo="text",
        hovertext=node_text,
        customdata=node_ids,
        marker=dict(size=node_sizes, color=node_colors, line=dict(width=2, color="#ffffff"), opacity=0.95),
        name="Customers",
    )

    traces = []
    if isinstance(cluster_map, dict) and cluster_map:
        halo_colors = []
        for n in ordered_ids:
            cid = int(cluster_map.get(str(n), -1))
            base = _cluster_highlight_color(cid)
            halo_colors.append(_hex_to_rgba(base, 0.24))

        halo_trace = go.Scatter(
            x=node_x,
            y=node_y,
            mode="markers",
            hoverinfo="skip",
            marker=dict(
                size=[s + 20 for s in node_sizes],
                color=halo_colors,
                line=dict(width=0),
                opacity=1.0,
            ),
            name="Network clusters",
        )
        traces.append(halo_trace)

    traces.append(node_trace)

    # Directed arrows as annotations (no overlapping line layer needed)
    node_score_map = {n: float(getattr(id_to_row[n], "final_node_risk_score", 0.0) or 0.0) for n in ordered_ids}
    annotations = []
    for e in edge_df.itertuples(index=False):
        u = str(e.originator_id)
        v = str(e.beneficiary_id)
        if u not in pos or v not in pos:
            continue

        u_cluster = int(cluster_map.get(u, -1)) if isinstance(cluster_map, dict) and cluster_map else -1
        v_cluster = int(cluster_map.get(v, -1)) if isinstance(cluster_map, dict) and cluster_map else -1
        is_active_edge = (
            active_cluster_id is None
            or (
                (u_cluster in (0, int(active_cluster_id)))
                and (v_cluster in (0, int(active_cluster_id)))
            )
        )

        x0, y0 = pos[u]
        x1, y1 = pos[v]
        dx, dy = x1 - x0, y1 - y0
        length = max((dx ** 2 + dy ** 2) ** 0.5, 1e-9)

        src_idx = ordered_ids.index(u)
        dst_idx = ordered_ids.index(v)
        src_size = node_sizes[src_idx]
        dst_size = node_sizes[dst_idx]
        src_offset = min(0.26, max(0.08, (src_size / 110.0)))
        dst_offset = min(0.28, max(0.09, (dst_size / 100.0)))

        ax = x0 + (dx / length) * src_offset
        ay = y0 + (dy / length) * src_offset
        x = x1 - (dx / length) * dst_offset
        y = y1 - (dy / length) * dst_offset

        edge_score = max(node_score_map.get(u, 0.0), node_score_map.get(v, 0.0))
        edge_color = edge_risk_color(edge_score) if is_active_edge else "#d4d8e0"
        edge_opacity = 0.85 if is_active_edge else 0.40
        edge_width = 1.8 if is_active_edge else 2.3
        annotations.append(dict(
            ax=ax, ay=ay,
            x=x, y=y,
            xref="x", yref="y", axref="x", ayref="y",
            showarrow=True, arrowhead=2, arrowsize=1.2,
            arrowwidth=edge_width, arrowcolor=edge_color, opacity=edge_opacity,
        ))

    fig = go.Figure(data=traces)
    fig.update_layout(
        height=640,
        margin=dict(l=10, r=10, t=10, b=10),
        showlegend=False,
        dragmode="pan",
        xaxis=dict(visible=False),
        yaxis=dict(visible=False),
        plot_bgcolor="#fafbff",
        paper_bgcolor="#fafbff",
        annotations=annotations,
        hoverlabel=dict(bgcolor="white", font_size=13, font_family="Helvetica"),
    )
    return fig


def render_color_legend() -> None:
    node_items = [
        ("#1f4aff", "Target Customer"),
        ("#d73027", "Sanctions Match"),
        ("#7a0177", "SAR Filed"),
        ("#f39c12", "PEP"),
        ("#00a6a6", "Exited Customer"),
        ("#4d4d4d", "Standard Node"),
    ]
    edge_items = [
        ("#b30000", "Very High Risk"),
        ("#c95f02", "High Risk"),
        ("#856404", "Medium Risk"),
        ("#2e7d32", "Low Risk"),
    ]
    parts = [
        '<div style="display:flex;flex-wrap:wrap;gap:14px;padding:8px 4px;'
        'border-top:1px solid #e0e4ef;margin-top:4px;">',
        '<span style="font-size:0.8rem;font-weight:600;color:#ffffff;align-self:center;">Node colours:</span>',
    ]
    for color, label in node_items:
        parts.append(
            f'<span style="display:inline-flex;align-items:center;gap:5px;">'
            f'<svg width="14" height="14"><circle cx="7" cy="7" r="6" fill="{color}" '
            f'stroke="#fff" stroke-width="1.5"/></svg>'
            f'<span style="font-size:0.8rem;color:#ffffff;">{label}</span></span>'
        )
    parts.append("</div>")
    parts.append(
        '<div style="display:flex;flex-wrap:wrap;gap:14px;padding:8px 4px;'
        'border-top:1px solid #e0e4ef;margin-top:4px;">'
        '<span style="font-size:0.8rem;font-weight:600;color:#ffffff;align-self:center;">Arrow/edge colours '
        '(risk band of the riskier endpoint, matching the risk summary):</span>'
    )
    for color, label in edge_items:
        parts.append(
            f'<span style="display:inline-flex;align-items:center;gap:5px;">'
            f'<svg width="20" height="10"><line x1="1" y1="5" x2="19" y2="5" stroke="{color}" '
            f'stroke-width="3"/></svg>'
            f'<span style="font-size:0.8rem;color:#ffffff;">{label}</span></span>'
        )
    parts.append(
        '<span style="font-size:0.78rem;color:#ffffff;align-self:center;margin-left:6px;">'
        '↗ Arrows = transaction direction (originator → beneficiary), attached to node boundary</span>'
    )
    parts.append("</div>")
    parts.append(
        '<div style="padding:6px 4px 2px;color:#ffffff;font-size:0.78rem;">'
        'In Full Network view, soft transparent halos indicate the split sub-network each node belongs to '
        '(target customer appears in all split networks).</div>'
    )
    st.markdown("".join(parts), unsafe_allow_html=True)


def render_network_score_explanation() -> None:
    with st.expander("ℹ️ How are Risk Scores calculated?", expanded=False):
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("""**Network Risk Score** (0–100)

| Component | Weight | What it measures |
|---|---|---|
| Theme Severity | 35% | Avg severity of all triggered themes |
| High-Risk Node Share | 20% | % of nodes scoring ≥ 70 |
| Value Factor | 15% | Log-scaled total flow value |
| Velocity Intensity | 10% | High-frequency burst patterns |
| Circularity Intensity | 10% | Closed-loop / layering patterns |
| Exposure Factor | 10% | Flagged-node count × value share |

Applied formula:

`Network Score = 0.35×Theme + 0.20×HighRiskShare + 0.15×Value + 0.10×Velocity + 0.10×Circularity + 0.10×Exposure`

Each component is normalized to 0–100 before weighting, then summed and capped at 100.
""")
        with col2:
            st.markdown("""**Node Risk Score** (effective range: 0–58.5)

| Component | Weight | What it measures |
|---|---|---|
| PageRank | 40% | Network influence relative to peers |
| Risk Flags | 30% | Sanctions+30, SAR+22, PEP+15, Exited+12 |
| Proximity | 20% | Hop-distance to nearest flagged node |
| Behaviour | 10% | Velocity bursts, structuring, pass-through |

Applied formula:

`Node Score = 0.40×PageRank + 0.30×Flags + 0.20×Proximity + 0.10×Behaviour`

Internal caps before weighting:

- `Flags = min(40, sanctions(30) + sar(22) + pep(15) + exited(12))`
- `Proximity = 20 (d<=1), 12 (d=2), 6 (d=3), else 0`
- `Behaviour = min(25, behavioural points)`

Maximum possible node score with current internal caps:

`0.40×100 + 0.30×40 + 0.20×20 + 0.10×25 = 58.5`

The weighted components shown in the breakdown are the actual contribution values after weight application.
""")


def resolve_effective_network_for_node(
    analyzer,
    node_rankings: pd.DataFrame,
    selected_network: str,
    selected_node: str,
    target_customer_id: str,
) -> str:
    selected_network_str = str(selected_network)
    selected_node_str = str(selected_node)
    target_str = str(target_customer_id)
    full_network_id = str(getattr(analyzer, "full_network_id", "Network-Full"))

    # Target-customer rule:
    # Full Network -> top network, Specific Network -> that network.
    if selected_node_str == target_str:
        if selected_network_str == full_network_id:
            if isinstance(analyzer.network_summary, pd.DataFrame) and not analyzer.network_summary.empty:
                return str(analyzer.network_summary.iloc[0].get("network_id", selected_network_str))
        return selected_network_str

    # Non-target fallback under Full Network focus.
    if selected_network_str == full_network_id and not node_rankings.empty:
        selected_node_rows = node_rankings[node_rankings["customer_id"].astype(str) == selected_node_str].copy()
        if not selected_node_rows.empty:
            return str(selected_node_rows.iloc[0].get("network_id", selected_network_str))

    return selected_network_str


def render_dynamic_score_breakdown(analyzer, outputs, selected_network: str, selected_node: str) -> None:
    node_rankings = outputs.get("node_rankings", pd.DataFrame())
    theme_triggers = outputs.get("theme_triggers", pd.DataFrame())
    target_cid = str(getattr(analyzer, "target_customer_id", ""))
    effective_network = resolve_effective_network_for_node(
        analyzer=analyzer,
        node_rankings=node_rankings,
        selected_network=selected_network,
        selected_node=selected_node,
        target_customer_id=target_cid,
    )
    is_target_node = str(selected_node) == target_cid

    with st.expander("🧮 Actual Values Used For Computation (For Demo purposes only, not for FCI)", expanded=True):
        net_col, node_col = st.columns(2)

        with net_col:
            st.markdown("**Network Risk Score (Selected Network)**")
            net_row_df = analyzer.network_summary[
                analyzer.network_summary["network_id"].astype(str) == str(effective_network)
            ]
            if net_row_df.empty:
                st.info("No network score data available for selected network.")
            else:
                net_row = net_row_df.iloc[0]
                total_value = float(net_row.get("total_amount_usd", 0.0) or 0.0)

                if is_target_node and str(selected_network) == str(getattr(analyzer, "full_network_id", "Network-Full")):
                    st.caption(f"Showing score inputs for top network {effective_network} for the target customer.")
                elif str(effective_network) != str(selected_network):
                    st.caption(f"Showing score inputs for {effective_network} because Full Network focus is selected.")

                net_nodes = node_rankings[node_rankings["network_id"].astype(str) == str(effective_network)].copy()
                net_themes = theme_triggers[theme_triggers["network_id"].astype(str) == str(effective_network)].copy()

                theme_severity = float(net_themes["severity_score"].mean()) if not net_themes.empty else 0.0
                high_risk_share = float((net_nodes["final_node_risk_score"] >= 70).mean() * 100) if not net_nodes.empty else 0.0
                value_factor = float(min(100.0, math.log1p(max(0.0, total_value)) * 8.0))

                velocity_rows = net_themes[net_themes["subtheme"].astype(str) == "High Velocity Transactions"]
                velocity_intensity = float(velocity_rows["severity_score"].mean()) if not velocity_rows.empty else 0.0

                circular_rows = net_themes[net_themes["subtheme"].astype(str) == "Closed Loop / Circular Flow"]
                circularity_intensity = float(circular_rows["severity_score"].mean()) if not circular_rows.empty else 0.0

                flagged_nodes = set()
                if not net_nodes.empty:
                    flagged_mask = (
                        net_nodes["sanctions_flag"].astype(bool)
                        | net_nodes["pep_flag"].astype(bool)
                        | net_nodes["sar_flag"].astype(bool)
                        | net_nodes["exited_flag"].astype(bool)
                    )
                    flagged_nodes = set(net_nodes.loc[flagged_mask, "customer_id"].astype(str).tolist())

                exposure_count = float(len(flagged_nodes))
                exposure_txn_value = 0.0
                txn_df = analyzer.network_txn.get(str(effective_network), pd.DataFrame()).copy()
                if not txn_df.empty and flagged_nodes:
                    impacted = txn_df[
                        txn_df["originator_id"].astype(str).isin(flagged_nodes)
                        | txn_df["beneficiary_id"].astype(str).isin(flagged_nodes)
                    ]
                    exposure_txn_value = float(impacted["amount_usd"].sum()) if not impacted.empty else 0.0

                exposure_factor = float(min(100.0, exposure_count * 8.0 + (exposure_txn_value / max(1.0, total_value)) * 60.0))

                comp_df = pd.DataFrame([
                    {"Component": "Theme Severity", "Raw Value": theme_severity, "Weight": 0.35},
                    {"Component": "High-Risk Node Share", "Raw Value": high_risk_share, "Weight": 0.20},
                    {"Component": "Value Factor", "Raw Value": value_factor, "Weight": 0.15},
                    {"Component": "Velocity Intensity", "Raw Value": velocity_intensity, "Weight": 0.10},
                    {"Component": "Circularity Intensity", "Raw Value": circularity_intensity, "Weight": 0.10},
                    {"Component": "Exposure Factor", "Raw Value": exposure_factor, "Weight": 0.10},
                ])
                comp_df["Weighted Contribution"] = comp_df["Raw Value"] * comp_df["Weight"]
                network_score_calc = float(min(100.0, comp_df["Weighted Contribution"].sum()))

                st.dataframe(
                    comp_df.assign(
                        **{
                            "Raw Value": comp_df["Raw Value"].map(lambda x: f"{x:.2f}"),
                            "Weight": comp_df["Weight"].map(lambda x: f"{x:.0%}"),
                            "Weighted Contribution": comp_df["Weighted Contribution"].map(lambda x: f"{x:.2f}"),
                        }
                    ),
                    use_container_width=True,
                )
                st.markdown(
                    f"**Resultant Network Risk Score:** `{network_score_calc:.2f}` "
                    f"(stored: `{float(net_row.get('network_risk_score', 0.0) or 0.0):.2f}`)"
                )

        with node_col:
            st.markdown("**Node Risk Score (Selected Node)**")
            node_row_df = node_rankings[
                (node_rankings["network_id"].astype(str) == str(effective_network))
                & (node_rankings["customer_id"].astype(str) == str(selected_node))
            ]

            if node_row_df.empty:
                st.info("No node score data available for selected node.")
            else:
                row = node_row_df.iloc[0]
                pagerank_c = float(row.get("pagerank_component", 0.0) or 0.0)
                flags_c = float(row.get("flags_component", 0.0) or 0.0)
                proximity_c = float(row.get("proximity_component", 0.0) or 0.0)
                behaviour_c = float(row.get("behaviour_component", 0.0) or 0.0)
                pagerank_raw = float(row.get("pagerank_raw_value", 0.0) or 0.0)
                pagerank_raw_max = float(row.get("pagerank_raw_max_in_network", 0.0) or 0.0)
                pagerank_norm = float(row.get("pagerank_normalized_score", 0.0) or 0.0)

                flags_raw = flags_c / 0.3 if 0.3 else 0.0
                proximity_raw = proximity_c / 0.2 if 0.2 else 0.0
                behaviour_raw = behaviour_c / 0.1 if 0.1 else 0.0

                node_comp_df = pd.DataFrame([
                    {
                        "Component": "PageRank",
                        "Weight": "40%",
                        "Weighted Contribution": pagerank_c,
                        "Raw Value": pagerank_norm,
                        "Raw Value Calculation": (
                            f"({pagerank_raw:.8f} / max_raw {pagerank_raw_max:.8f}) x 100 = {pagerank_norm:.2f}"
                        ),
                    },
                    {
                        "Component": "Flags",
                        "Weight": "30%",
                        "Weighted Contribution": flags_c,
                        "Raw Value": flags_raw,
                        "Raw Value Calculation": f"{flags_c:.2f} / 0.30 = {flags_raw:.2f}",
                    },
                    {
                        "Component": "Proximity",
                        "Weight": "20%",
                        "Weighted Contribution": proximity_c,
                        "Raw Value": proximity_raw,
                        "Raw Value Calculation": f"{proximity_c:.2f} / 0.20 = {proximity_raw:.2f}",
                    },
                    {
                        "Component": "Behaviour",
                        "Weight": "10%",
                        "Weighted Contribution": behaviour_c,
                        "Raw Value": behaviour_raw,
                        "Raw Value Calculation": f"{behaviour_c:.2f} / 0.10 = {behaviour_raw:.2f}",
                    },
                ])
                node_score_calc = float(min(NODE_SCORE_MAX, node_comp_df["Weighted Contribution"].sum()))

                internal_inputs_df = pd.DataFrame([
                    {
                        "Internal Input": "PageRank (raw)",
                        "Estimated Value": pagerank_raw,
                        "Cap": "n/a",
                        "How calculated": "Direct NetworkX pagerank output before normalization",
                    },
                    {
                        "Internal Input": "PageRank (normalized)",
                        "Estimated Value": pagerank_norm,
                        "Cap": "100",
                        "How calculated": f"({pagerank_raw:.8f} / max_raw {pagerank_raw_max:.8f}) x 100",
                    },
                    {
                        "Internal Input": "Flags (post-cap)",
                        "Estimated Value": flags_raw,
                        "Cap": "40",
                        "How calculated": f"{flags_c:.2f} / 0.30",
                    },
                    {
                        "Internal Input": "Proximity (post-bucket)",
                        "Estimated Value": proximity_raw,
                        "Cap": "20",
                        "How calculated": f"{proximity_c:.2f} / 0.20",
                    },
                    {
                        "Internal Input": "Behaviour (post-cap)",
                        "Estimated Value": behaviour_raw,
                        "Cap": "25",
                        "How calculated": f"{behaviour_c:.2f} / 0.10",
                    },
                ])

                st.dataframe(
                    node_comp_df.assign(
                        **{
                            "Raw Value": node_comp_df["Raw Value"].map(lambda x: f"{x:.2f}"),
                            "Weighted Contribution": node_comp_df["Weighted Contribution"].map(lambda x: f"{x:.2f}")
                        }
                    ),
                    use_container_width=True,
                )
                st.dataframe(
                    internal_inputs_df.assign(
                        **{"Estimated Value": internal_inputs_df["Estimated Value"].map(lambda x: f"{x:.2f}")}
                    ),
                    use_container_width=True,
                )
                st.caption(
                    f"Internal equation for selected node: {pagerank_c:.2f} + {flags_c:.2f} + {proximity_c:.2f} + {behaviour_c:.2f} = {node_score_calc:.2f} "
                    f"(max possible with current caps: {NODE_SCORE_MAX:.1f})."
                )
                st.markdown(
                    f"**Resultant Node Risk Score ({selected_node}):** `{node_score_calc:.2f}` "
                    f"(stored: `{float(row.get('final_node_risk_score', 0.0) or 0.0):.2f}`)"
                )


def render_case_snapshot(
    analyzer,
    selected_node: str,
    target_customer_id: str,
    selected_network: str,
    full_network_id: str,
    txn_scope_network: str,
    node_df: pd.DataFrame,
    node_rankings: pd.DataFrame,
    node_tx: pd.DataFrame,
    kyc: dict,
    scoped_network_tx_export: pd.DataFrame,
    all_networks_tx_export: pd.DataFrame,
) -> None:
    cid = str(selected_node)
    target_cid = str(target_customer_id)
    effective_network = str(selected_network)
    if cid == target_cid:
        if str(selected_network) == str(full_network_id):
            if isinstance(analyzer.network_summary, pd.DataFrame) and not analyzer.network_summary.empty:
                effective_network = str(analyzer.network_summary.iloc[0].get("network_id", selected_network))
        else:
            effective_network = str(selected_network)
    elif str(selected_network) == str(full_network_id):
        node_matches = node_rankings[node_rankings["customer_id"].astype(str) == cid]
        if not node_matches.empty:
            effective_network = str(node_matches.iloc[0].get("network_id", selected_network))

    rank_match = node_rankings[
        (node_rankings["network_id"].astype(str) == str(effective_network))
        & (node_rankings["customer_id"].astype(str) == cid)
    ]

    risk_score = pagerank_c = flags_c = prox_c = beh_c = 0.0
    key_reasons = ""
    sanctions = pep = sar = exited = False
    if not rank_match.empty:
        row = rank_match.iloc[0]
        risk_score = float(row.get("final_node_risk_score", 0.0) or 0.0)
        pagerank_c = float(row.get("pagerank_component", 0.0) or 0.0)
        flags_c = float(row.get("flags_component", 0.0) or 0.0)
        prox_c = float(row.get("proximity_component", 0.0) or 0.0)
        beh_c = float(row.get("behaviour_component", 0.0) or 0.0)
        key_reasons = str(row.get("key_reasons", "") or "")
        sanctions = bool(row.get("sanctions_flag", False))
        pep = bool(row.get("pep_flag", False))
        sar = bool(row.get("sar_flag", False))
        exited = bool(row.get("exited_flag", False))

    risk_band, band_color = node_risk_band_from_internal(risk_score)

    left_col, right_col = st.columns(2)

    with left_col:
        with st.container(border=True):
            card_title("🪪", "KYC Details")
            f1, f2, f3 = st.columns(3)
            f1.markdown(f"**Name**\n\n{kyc.get('customer_name', '—') or '—'}")
            f2.markdown(f"**Country**\n\n{kyc.get('country', '—') or '—'}")
            f3.markdown(f"**Entity Type**\n\n{kyc.get('entity_type', '—') or '—'}")
            f4, f5, f6 = st.columns(3)
            f4.markdown(f"**Gender**\n\n{kyc.get('gender', '—') or '—'}")
            f5.markdown(f"**DOB / Incorp**\n\n{kyc.get('date_of_birth_or_incorp', '—') or '—'}")
            f6.markdown(f"**Exit Date**\n\n{kyc.get('exit_date', '—') or '—'}")
            flags_html = ""
            if sanctions:
                flags_html += '<span class="badge badge-red">⛔ Sanctions</span>'
            if pep:
                flags_html += '<span class="badge badge-orange">🎖️ PEP</span>'
            if sar:
                flags_html += '<span class="badge badge-red">📋 SAR</span>'
            if exited:
                flags_html += '<span class="badge badge-purple">🚪 Exited</span>'
            if not flags_html:
                flags_html = '<span class="badge badge-blue">✅ No Flags</span>'
            st.markdown(f"**Risk Flags:** {flags_html}", unsafe_allow_html=True)

    with right_col:
        with st.container(border=True):
            card_title("🔎", "Connected Parties With High DRA Score")
            dra_threshold = st.number_input(
                "Simulate with DRA Score",
                min_value=0.0,
                max_value=5000.0,
                value=750.0,
                step=10.0,
                key=f"connected_party_dra_threshold_{target_cid}",
                help="Shows connected parties for the targeted customer across all discovered networks where DRA score is above this value.",
            )

            connected_nodes_df = pd.DataFrame()
            full_details = getattr(analyzer, "full_network_node_details", pd.DataFrame())
            if isinstance(full_details, pd.DataFrame) and not full_details.empty:
                connected_nodes_df = full_details.copy()
            else:
                profiles = getattr(analyzer, "node_profiles", pd.DataFrame())
                full_nodes = set([str(x) for x in getattr(analyzer, "full_network_nodes", set())])
                if isinstance(profiles, pd.DataFrame) and not profiles.empty and full_nodes:
                    connected_nodes_df = profiles[
                        profiles["customer_id"].astype(str).isin(full_nodes)
                    ][["customer_id", "customer_name", "dra_score"]].copy()

            if connected_nodes_df.empty:
                st.info("No connected-party data available for the targeted customer in the discovered networks.")
            else:
                connected_nodes_df["customer_id"] = connected_nodes_df["customer_id"].astype(str)
                connected_nodes_df["customer_name"] = connected_nodes_df["customer_name"].fillna("").astype(str)
                connected_nodes_df["dra_score"] = pd.to_numeric(connected_nodes_df["dra_score"], errors="coerce").fillna(0.0)

                network_id_map = {}
                if isinstance(node_rankings, pd.DataFrame) and not node_rankings.empty:
                    rank_lookup = node_rankings[["customer_id", "network_id"]].copy()
                    rank_lookup["customer_id"] = rank_lookup["customer_id"].astype(str)
                    rank_lookup["network_id"] = rank_lookup["network_id"].astype(str)
                    network_id_map = (
                        rank_lookup
                        .drop_duplicates(subset=["customer_id"], keep="first")
                        .set_index("customer_id")["network_id"]
                        .to_dict()
                    )
                connected_nodes_df["network_id"] = connected_nodes_df["customer_id"].map(network_id_map).fillna("Unknown")

                high_dra_parties = connected_nodes_df[
                    (connected_nodes_df["customer_id"] != target_cid)
                    & (connected_nodes_df["dra_score"] > float(dra_threshold))
                ][["customer_id", "customer_name", "dra_score", "network_id"]].drop_duplicates(subset=["customer_id"]).copy()

                high_dra_parties = high_dra_parties.sort_values("dra_score", ascending=False).reset_index(drop=True)

                st.caption(
                    f"Targeted customer: {target_cid}. Search scope: all discovered networks connected to the target. "
                    f"Threshold: DRA > {float(dra_threshold):.1f}."
                )

                if high_dra_parties.empty:
                    st.success("No connected parties found above the configured DRA threshold.")
                else:
                    st.dataframe(
                        high_dra_parties.rename(
                            columns={
                                "network_id": "Network ID",
                                "customer_id": "Customer ID",
                                "customer_name": "Customer Name",
                                "dra_score": "DRA Score",
                            }
                        ),
                        use_container_width=True,
                        hide_index=True,
                    )
                    st.caption(f"Found {len(high_dra_parties)} connected parties above threshold.")

    with st.container(border=True):
        node_tx_display = node_tx.copy()
        if not node_tx_display.empty and "network_id" not in node_tx_display.columns:
            if (
                isinstance(scoped_network_tx_export, pd.DataFrame)
                and not scoped_network_tx_export.empty
                and "transaction_id" in node_tx_display.columns
                and "transaction_id" in scoped_network_tx_export.columns
                and "network_id" in scoped_network_tx_export.columns
            ):
                network_lookup = scoped_network_tx_export[["transaction_id", "network_id"]].drop_duplicates("transaction_id")
                node_tx_display = node_tx_display.merge(network_lookup, on="transaction_id", how="left")
                node_tx_display["network_id"] = node_tx_display["network_id"].fillna(str(txn_scope_network))
            else:
                node_tx_display.insert(0, "network_id", str(txn_scope_network))

            cols = ["network_id"] + [c for c in node_tx_display.columns if c != "network_id"]
            node_tx_display = node_tx_display[cols]

        card_title("💳", f"Linked Transactions ({len(node_tx)} total)")
        st.dataframe(node_tx_display, use_container_width=True)
        dcol1, dcol2 = st.columns(2)
        with dcol1:
            st.download_button(
                "📥 Download All Transactions (Selected Network Focus)",
                data=to_csv_bytes(scoped_network_tx_export),
                file_name=f"transactions_{str(txn_scope_network)}.csv",
                mime="text/csv",
                key=f"download_selected_network_tx_{str(txn_scope_network)}",
            )
        with dcol2:
            st.download_button(
                "📥 Download All Transactions (All Networks)",
                data=to_csv_bytes(all_networks_tx_export),
                file_name="transactions_all_networks.csv",
                mime="text/csv",
                key="download_all_networks_tx",
            )


def render_executive_summary_cards(analyzer, outputs) -> None:
    summary = analyzer.network_summary
    if summary.empty:
        st.info("No networks found for the selected target in the configured lookback window.")
        return

    top_net = summary.iloc[0]
    node_table = outputs.get("node_rankings", pd.DataFrame())
    total_flagged = 0
    if not node_table.empty:
        flag_cols = [c for c in ["sanctions_flag", "pep_flag", "sar_flag", "exited_flag"] if c in node_table.columns]
        if flag_cols:
            total_flagged = int(node_table[flag_cols].any(axis=1).sum())

    top_network_score = float(top_net["network_risk_score"])
    top_nodes = int(top_net["nodes"])

    top_network_id = str(top_net["network_id"])
    top_node_total_usd = 0.0
    top_node_id = "N/A"
    if not node_table.empty and "network_id" in node_table.columns and "final_node_risk_score" in node_table.columns:
        top_net_nodes = node_table[node_table["network_id"].astype(str) == top_network_id]
        if not top_net_nodes.empty:
            top_node_row = top_net_nodes.sort_values("final_node_risk_score", ascending=False).iloc[0]
            top_node_id = str(top_node_row.get("customer_id", "N/A") or "N/A")
            top_node_tx = analyzer.get_node_transactions(top_network_id, top_node_id)
            if isinstance(top_node_tx, pd.DataFrame) and not top_node_tx.empty and "amount_usd" in top_node_tx.columns:
                top_node_total_usd = float(pd.to_numeric(top_node_tx["amount_usd"], errors="coerce").fillna(0.0).sum())

    # Show the selected target customer's actual DRA score from node profiles.
    selected_customer_dra = 0.0
    try:
        profiles = getattr(analyzer, "node_profiles", pd.DataFrame())
        if isinstance(profiles, pd.DataFrame) and not profiles.empty and "dra_score" in profiles.columns:
            target_rows = profiles[profiles["customer_id"].astype(str) == str(analyzer.target_customer_id)]
            if not target_rows.empty:
                selected_customer_dra = float(pd.to_numeric(target_rows["dra_score"], errors="coerce").fillna(0.0).max())
    except Exception:
        selected_customer_dra = 0.0

    def _compact_usd(value: float) -> str:
        v = float(value or 0.0)
        av = abs(v)
        if av >= 1_000_000_000:
            return f"${v/1_000_000_000:.1f}B"
        if av >= 1_000_000:
            return f"${v/1_000_000:.1f}M"
        if av >= 1_000:
            return f"${v/1_000:.1f}K"
        return f"${v:,.0f}"

    m1, m2, m3, m4, m5, m6 = st.columns(6)
    m1.metric("Networks Found", int(len(summary)))
    m2.metric("Top Network Score", f"{top_network_score:.0f}", help="Score is on a 0-100 scale")
    m3.metric("Nodes (Top Network)", top_nodes)
    m4.metric("Flagged Nodes (All)", total_flagged)
    m5.metric(
        "Top Node USD",
        _compact_usd(top_node_total_usd),
        help=f"Total amount_usd linked to the highest-score node in {top_network_id} (customer: {top_node_id})",
    )
    m6.metric("DRA Score", f"{selected_customer_dra:.1f}", help="Selected customer DRA score from profile fields in the loaded dataset")

    st.markdown(
        f"**Lookback:** {analyzer.window_label} &nbsp;|&nbsp; **Target:** `{analyzer.target_customer_id}` "
        f"&nbsp;|&nbsp; **Top network:** {top_net['network_id']} ({top_net['flagged_nodes_counts']})",
        unsafe_allow_html=True,
    )

    top_themes = analyzer.theme_log[
        analyzer.theme_log["network_id"] == top_net["network_id"]
    ].sort_values("severity_score", ascending=False).head(5)

    # Removed Top triggered themes heading from Streamlit UI; keeping code commented out for later reuse.
    # if not top_themes.empty:
    #     st.markdown("**Top triggered themes:**")
    #     for r in top_themes.itertuples(index=False):
    #         clr = "#b30000" if r.severity_score >= 70 else ("#c95f02" if r.severity_score >= 40 else "#1f77b4")
    #         st.markdown(
    #             f'<div style="margin:4px 0;padding:6px 10px;background:#fff;border-radius:6px;border-left:4px solid {clr};">'
    #             f'<span style="color:{clr};font-weight:700;">{r.subtheme}</span>'
    #             f'<span style="color:#888;font-size:0.8rem;"> ({r.theme})</span>'
    #             f'</div>',
    #             unsafe_allow_html=True,
    #         )


def build_selected_customer_genai_summary(
    selected_network: str,
    selected_customer_id: str,
    node_df: pd.DataFrame,
    node_rankings: pd.DataFrame,
    node_tx: pd.DataFrame,
) -> str:
    cid = str(selected_customer_id)
    node_match = node_df[node_df["customer_id"].astype(str) == cid]
    rank_match = node_rankings[
        (node_rankings["network_id"].astype(str) == str(selected_network))
        & (node_rankings["customer_id"].astype(str) == cid)
    ]

    if node_match.empty and rank_match.empty:
        return "No summary available for the selected customer in this network."

    name = cid
    country = ""
    entity_type = ""
    risk_score = 0.0
    reasons = ""
    sanctions = pep = sar = exited = False

    if not node_match.empty:
        n = node_match.iloc[0]
        name = str(n.get("customer_name", cid) or cid)
        country = str(n.get("country", "") or "")
        entity_type = str(n.get("entity_type", "") or "")
        risk_score = float(n.get("final_node_risk_score", 0.0) or 0.0)
        sanctions = bool(n.get("sanctions_flag", False))
        pep = bool(n.get("pep_flag", False))
        sar = bool(n.get("sar_flag", False))
        exited = bool(n.get("exited_flag", False))

    if not rank_match.empty:
        r = rank_match.iloc[0]
        risk_score = float(r.get("final_node_risk_score", risk_score) or risk_score)
        reasons = str(r.get("key_reasons", "") or "")

    txn_count = int(len(node_tx))
    total_value = float(node_tx["amount_usd"].sum()) if txn_count else 0.0
    avg_value = float(node_tx["amount_usd"].mean()) if txn_count else 0.0

    outgoing = node_tx[node_tx["originator_id"].astype(str) == cid]
    incoming = node_tx[node_tx["beneficiary_id"].astype(str) == cid]
    outgoing_count = int(len(outgoing))
    incoming_count = int(len(incoming))

    cross_border_ratio = 0.0
    if txn_count:
        cross_border_ratio = float(
            (node_tx["originator_id"].astype(str) != node_tx["beneficiary_id"].astype(str)).mean()
        )

    cp_series = pd.concat(
        [
            outgoing["beneficiary_id"].astype(str),
            incoming["originator_id"].astype(str),
        ],
        ignore_index=True,
    )
    top_counterparties = cp_series.value_counts().head(3)
    top_cp_text = ", ".join([f"{k} ({int(v)} txns)" for k, v in top_counterparties.items()]) if not top_counterparties.empty else "none"

    top_scenarios = (
        node_tx["scenario_tag"].astype(str).replace("nan", "").replace("None", "").replace("", pd.NA).dropna().value_counts().head(3)
        if txn_count
        else pd.Series(dtype=int)
    )
    top_scenarios_text = ", ".join([f"{k} ({int(v)})" for k, v in top_scenarios.items()]) if not top_scenarios.empty else "none"

    risk_band, _ = node_risk_band_from_internal(risk_score)
    risk_band = risk_band.replace("🔴 ", "").replace("🟠 ", "").replace("🟡 ", "").replace("🟢 ", "")

    flag_labels = []
    if sanctions:
        flag_labels.append("Sanctions")
    if pep:
        flag_labels.append("PEP")
    if sar:
        flag_labels.append("SAR")
    if exited:
        flag_labels.append("Exited")

    lines = []
    lines.append(f"Customer {cid} ({name}) is assessed as {risk_band} risk with score {risk_score:.2f} in {selected_network}.")
    lines.append(
        f"Profile context: country={country or 'unknown'}, entity_type={entity_type or 'unknown'}, active risk flags={', '.join(flag_labels) if flag_labels else 'none'}."
    )
    lines.append(
        f"Observed activity for this network: {txn_count} linked transactions totaling USD {total_value:,.2f} (avg USD {avg_value:,.2f}), with {outgoing_count} outgoing and {incoming_count} incoming flows."
    )
    lines.append(f"Most frequent counterparties: {top_cp_text}.")
    lines.append(f"Top scenario tags around this customer: {top_scenarios_text}.")
    if reasons and reasons.strip():
        lines.append(f"Primary model reasons: {reasons.replace('|', ', ')}.")
    lines.append(
        "Suggested investigator focus: validate counterparties and transaction narratives for the latest high-value or high-frequency flows tied to this customer."
    )
    return "\n\n".join(lines)


def build_investigator_target_summary(
    analyzer,
    outputs,
    selected_network: str,
    target_customer_id: str,
    node_df: pd.DataFrame,
    target_tx: pd.DataFrame,
    kyc: dict,
):
    cid = str(target_customer_id)
    node_rankings = outputs.get("node_rankings", pd.DataFrame())
    effective_network = resolve_effective_network_for_node(
        analyzer=analyzer,
        node_rankings=node_rankings,
        selected_network=selected_network,
        selected_node=cid,
        target_customer_id=cid,
    )
    rank_row = node_rankings[
        (node_rankings["network_id"].astype(str) == str(effective_network))
        & (node_rankings["customer_id"].astype(str) == cid)
    ]
    rank = rank_row.iloc[0] if not rank_row.empty else None

    network_row_df = analyzer.network_summary[
        analyzer.network_summary["network_id"].astype(str) == str(effective_network)
    ]
    network_row = network_row_df.iloc[0] if not network_row_df.empty else None

    themes_df = analyzer.theme_log[
        analyzer.theme_log["network_id"].astype(str) == str(effective_network)
    ].sort_values("severity_score", ascending=False).head(5)

    txn_count = int(len(target_tx))
    total_value = float(target_tx["amount_usd"].sum()) if txn_count else 0.0
    avg_value = float(target_tx["amount_usd"].mean()) if txn_count else 0.0
    outgoing_count = int((target_tx["originator_id"].astype(str) == cid).sum()) if txn_count else 0
    incoming_count = int((target_tx["beneficiary_id"].astype(str) == cid).sum()) if txn_count else 0

    node_score = float(rank.get("final_node_risk_score", 0.0) or 0.0) if rank is not None else 0.0
    network_score = float(network_row.get("network_risk_score", 0.0) or 0.0) if network_row is not None else 0.0
    risk_band, _ = node_risk_band_from_internal(node_score)
    risk_band = risk_band.replace("🔴 ", "").replace("🟠 ", "").replace("🟡 ", "").replace("🟢 ", "")

    pagerank_c = float(rank.get("pagerank_component", 0.0) or 0.0) if rank is not None else 0.0
    flags_c = float(rank.get("flags_component", 0.0) or 0.0) if rank is not None else 0.0
    prox_c = float(rank.get("proximity_component", 0.0) or 0.0) if rank is not None else 0.0
    beh_c = float(rank.get("behaviour_component", 0.0) or 0.0) if rank is not None else 0.0
    drivers = sorted(
        [
            ("pagerank influence", pagerank_c),
            ("direct risk flags", flags_c),
            ("proximity to flagged entities", prox_c),
            ("behavioural patterns", beh_c),
        ],
        key=lambda x: x[1],
        reverse=True,
    )
    top_driver_text = ", ".join([f"{k} ({v:.1f})" for k, v in drivers[:3]])

    flag_text = []
    if rank is not None:
        if bool(rank.get("sanctions_flag", False)):
            flag_text.append("Sanctions")
        if bool(rank.get("sar_flag", False)):
            flag_text.append("SAR")
        if bool(rank.get("pep_flag", False)):
            flag_text.append("PEP")
        if bool(rank.get("exited_flag", False)):
            flag_text.append("Exited")

    counterparties = pd.Series(dtype=str)
    if txn_count:
        outgoing = target_tx[target_tx["originator_id"].astype(str) == cid]
        incoming = target_tx[target_tx["beneficiary_id"].astype(str) == cid]
        counterparties = pd.concat(
            [outgoing["beneficiary_id"].astype(str), incoming["originator_id"].astype(str)],
            ignore_index=True,
        )
    top_counterparties = counterparties.value_counts().head(3)
    top_counterparty_text = ", ".join([f"{k} ({int(v)} txns)" for k, v in top_counterparties.items()]) if not top_counterparties.empty else "none"

    top_scenarios = (
        target_tx["scenario_tag"].astype(str).replace("nan", "").replace("None", "").replace("", pd.NA).dropna().value_counts().head(3)
        if txn_count
        else pd.Series(dtype=int)
    )
    top_scenario_text = ", ".join([f"{k} ({int(v)})" for k, v in top_scenarios.items()]) if not top_scenarios.empty else "none"

    top_theme_text = "; ".join(
        [f"{r.subtheme} severity {float(r.severity_score):.1f}" for r in themes_df.itertuples(index=False)]
    ) if not themes_df.empty else "none"

    customer_name = str(kyc.get("customer_name", "") or "")
    network_count = int(len(analyzer.network_summary))
    networks_word = "network" if network_count == 1 else "networks"

    summary = "\n\n".join([
        "EXECUTIVE SUMMARY\n"
        f"Target customer {cid} ({customer_name or 'name unavailable'}) is assessed as {risk_band} risk in {effective_network}, "
        f"with node score {node_score:.2f} and selected-network score {network_score:.2f}. "
        f"The analysis covered {analyzer.window_label} and discovered {network_count} {networks_word} around the target. "
        f"Within the selected network, the target has {txn_count} linked transactions totaling USD {total_value:,.2f}; activity is split across {outgoing_count} outgoing and {incoming_count} incoming flows. "
        "The summary should be treated as an investigator triage narrative based on available transaction-network outputs.",
        "KEY FINDINGS\n"
        f"- Primary score drivers are {top_driver_text}.\n"
        f"- Direct risk flags on the target: {', '.join(flag_text) if flag_text else 'none recorded in the provided fields'}.\n"
        f"- Top network themes: {top_theme_text}.\n"
        f"- Frequent counterparties: {top_counterparty_text}.\n"
        f"- Scenario tags around the target: {top_scenario_text}.",
        "EVIDENCE BASIS\n"
        "This summary uses the analyzer network ranking, node ranking, theme trigger log, target KYC profile, and target-linked transaction table generated from the local dataset.",
        "LIMITATIONS\n"
        "The summary is limited to the provided transaction schema and calculated risk indicators; it does not include external KYC files, adverse media, sanctions-screening evidence, or investigator disposition notes.",
    ])

    actions = [
        "Review the target customer's highest-value and most recent transactions with supporting narratives and counterparty purpose.",
        "Validate frequent counterparties against expected customer activity, ownership links, and first-hop risk exposure.",
        "Inspect the strongest network themes and example transaction IDs before deciding whether escalation is warranted.",
    ]
    if flag_text:
        actions.append("Refresh screening and document whether any sanctions, PEP, SAR, or exited-customer indicators are true matches.")
    if "High Velocity Transactions" in top_theme_text:
        actions.append("Reconstruct timestamp-level sequences for high-velocity windows involving the target or immediate counterparties.")
    if "Closed Loop / Circular Flow" in top_theme_text:
        actions.append("Trace circular or return-flow paths to confirm whether there is an economic rationale for the movement of funds.")
    actions.append("Prepare a concise case note with evidence, unresolved questions, and recommended disposition.")

    return summary, actions[:6], {
        "target_customer_id": cid,
        "customer_name": customer_name,
        "selected_network": effective_network,
        "window_label": analyzer.window_label,
        "network_count": network_count,
        "network_score": network_score,
        "node_score": node_score,
        "risk_band": risk_band,
        "txn_count": txn_count,
        "total_value": total_value,
        "avg_value": avg_value,
        "outgoing_count": outgoing_count,
        "incoming_count": incoming_count,
        "top_driver_text": top_driver_text,
        "flag_text": flag_text,
        "top_theme_text": top_theme_text,
        "top_counterparty_text": top_counterparty_text,
        "top_scenario_text": top_scenario_text,
    }


def build_ai_customer_brief(
    analyzer,
    outputs,
    selected_network: str,
    selected_node: str,
    node_df: pd.DataFrame,
    node_tx: pd.DataFrame,
    fallback_summary: str,
    fallback_actions,
):
    node_rankings = outputs.get("node_rankings", pd.DataFrame())
    rank_row = node_rankings[
        (node_rankings["network_id"].astype(str) == str(selected_network))
        & (node_rankings["customer_id"].astype(str) == str(selected_node))
    ]
    rank = rank_row.iloc[0] if not rank_row.empty else None

    network_row_df = analyzer.network_summary[
        analyzer.network_summary["network_id"].astype(str) == str(selected_network)
    ]
    network_row = network_row_df.iloc[0] if not network_row_df.empty else None

    top_themes_df = analyzer.theme_log[
        analyzer.theme_log["network_id"].astype(str) == str(selected_network)
    ].sort_values("severity_score", ascending=False).head(5)

    txn_count = int(len(node_tx))
    total_value = float(node_tx["amount_usd"].sum()) if txn_count else 0.0
    avg_value = float(node_tx["amount_usd"].mean()) if txn_count else 0.0
    outgoing_count = int((node_tx["originator_id"].astype(str) == str(selected_node)).sum()) if txn_count else 0
    incoming_count = int((node_tx["beneficiary_id"].astype(str) == str(selected_node)).sum()) if txn_count else 0
    sanctions = bool(rank["sanctions_flag"]) if rank is not None and "sanctions_flag" in rank else False
    pep = bool(rank["pep_flag"]) if rank is not None and "pep_flag" in rank else False
    sar = bool(rank["sar_flag"]) if rank is not None and "sar_flag" in rank else False
    exited = bool(rank["exited_flag"]) if rank is not None and "exited_flag" in rank else False
    node_score = float(rank.get("final_node_risk_score", 0.0) or 0.0) if rank is not None else 0.0
    pagerank_c = float(rank.get("pagerank_component", 0.0) or 0.0) if rank is not None else 0.0
    flags_c = float(rank.get("flags_component", 0.0) or 0.0) if rank is not None else 0.0
    prox_c = float(rank.get("proximity_component", 0.0) or 0.0) if rank is not None else 0.0
    beh_c = float(rank.get("behaviour_component", 0.0) or 0.0) if rank is not None else 0.0

    network_score = float(network_row.get("network_risk_score", 0.0) or 0.0) if network_row is not None else 0.0
    theme_names = [str(r.subtheme) for r in top_themes_df.itertuples(index=False)]

    direction = "balanced"
    if outgoing_count > incoming_count * 1.4:
        direction = "outgoing-heavy"
    elif incoming_count > outgoing_count * 1.4:
        direction = "incoming-heavy"

    drivers = [
        ("pagerank influence", pagerank_c),
        ("direct risk flags", flags_c),
        ("proximity to flagged entities", prox_c),
        ("behavioural patterns", beh_c),
    ]
    drivers = sorted(drivers, key=lambda x: x[1], reverse=True)
    top_driver_text = ", ".join([f"{k} ({v:.1f})" for k, v in drivers[:3]])

    flag_text = []
    if sanctions:
        flag_text.append("Sanctions")
    if sar:
        flag_text.append("SAR")
    if pep:
        flag_text.append("PEP")
    if exited:
        flag_text.append("Exited")

    summary_lines = [
        f"Customer {selected_node} is alerted in network {selected_network} with node risk score {node_score:.2f} (effective max {NODE_SCORE_MAX:.1f}) while network risk is {network_score:.2f}.",
        f"Observed activity includes {txn_count} linked transactions totaling USD {total_value:,.2f} (avg USD {avg_value:,.2f}), with {outgoing_count} outgoing and {incoming_count} incoming flows ({direction}).",
        f"Primary score drivers are {top_driver_text}.",
    ]
    if flag_text:
        summary_lines.append(f"Direct risk flags present: {', '.join(flag_text)}.")
    else:
        summary_lines.append("No direct sanctions/PEP/SAR/exited flags on this node; alert is likely driven by network influence, proximity, and behaviour signals.")
    if theme_names:
        summary_lines.append(f"Top network patterns around this customer: {', '.join(theme_names[:4])}.")

    actions = []
    actions.append("Validate end-to-end transaction purpose for the top 5 highest-value and latest transactions involving this customer.")
    if direction == "outgoing-heavy":
        actions.append("Prioritize source-of-funds and beneficiary due diligence for dominant outgoing flows, including supporting invoices/contracts.")
    elif direction == "incoming-heavy":
        actions.append("Prioritize source-of-wealth validation for major incoming counterparties and verify business rationale for incoming concentration.")
    else:
        actions.append("Review both inbound and outbound counterparties for consistency with stated customer profile and expected activity.")

    if sanctions or sar or pep:
        actions.append("Run immediate refreshed screening and adverse media checks for customer and first-hop counterparties; document any true matches.")
    if "High Velocity Transactions" in theme_names:
        actions.append("Perform timestamp-level velocity analysis to confirm burst behavior and identify coordinated transaction chains.")
    if "Closed Loop / Circular Flow" in theme_names:
        actions.append("Trace potential circular fund movement across hops and reconcile economic rationale for looped transfers.")
    if beh_c >= max(pagerank_c, flags_c, prox_c):
        actions.append("Deep-dive behavioural triggers (structuring, pass-through, fan-out) and test against historical customer baseline.")
    if prox_c >= max(pagerank_c, flags_c, beh_c):
        actions.append("Map 1-2 hop connections to flagged entities and collect evidence explaining proximity risk propagation.")

    actions.append("Compile investigation chronology, evidence pack, and disposition recommendation (monitor / EDD / escalation / SAR review).")

    dedup_actions = []
    seen = set()
    for action in actions:
        key = action.strip().lower()
        if key not in seen:
            seen.add(key)
            dedup_actions.append(action)

    if len(dedup_actions) < 4:
        dedup_actions.extend(list(fallback_actions))

    return "\n\n".join(summary_lines) if summary_lines else fallback_summary, dedup_actions[:6], "Local investigator agent used (Copilot model is not directly callable from Streamlit runtime)."


st.markdown("""
<div class="app-header">
  <h1>🔍 Atlas Detection Lab</h1>
  <p>AI-Enabled Network Intelligence & Recommendation Engine</p>
</div>
""", unsafe_allow_html=True)

with st.sidebar:
    st.markdown("### ⚙️ Configuration")
    use_all_dates = st.checkbox("Use all dates", value=False)
    lookback_days = st.number_input("Lookback days (default 180)", min_value=30, max_value=3650, value=180, step=30)
    n_hops = st.number_input("Number of Counterparty Rings", min_value=1, max_value=6, value=2, step=1)
    ranking_mode = "value-weighted"
    use_ai_agent = st.checkbox("Use AI Summary Generator", value=True)
    ai_provider = st.selectbox(
        "AI provider",
        ["OpenAI", "Gemini", "Groq", "Ollama (Local)", "Deterministic"],
        index=0,
    )
    if ai_provider == "OpenAI":
        ai_model = st.selectbox(
            "OpenAI model",
            ["gpt-5.5", "gpt-4o", "gpt-4o-mini", "gpt-3.5-turbo"],
            index=2,
        )
    elif ai_provider == "Gemini":
        ai_model = st.selectbox(
            "Gemini model",
            ["gemini-2.5-flash", "gemini-2.5-pro", "gemini-2.5-flash-lite"],
            index=0,
        )
    elif ai_provider == "Groq":
        ai_model = st.selectbox(
            "Groq model",
            ["qwen/qwen3-32b", "llama-3.3-70b-versatile", "llama-3.1-8b-instant", "openai/gpt-oss-120b", "openai/gpt-oss-20b"],
            index=0,
        )
    else:
        ai_model = ""
    ollama_model = st.text_input("Ollama model", value="llama3.1:8b") if ai_provider == "Ollama (Local)" else "llama3.1:8b"

    if ai_provider == "Gemini" and not GEMINI_API_KEY:
        st.warning("Gemini is selected but GEMINI_API_KEY or GOOGLE_API_KEY is not configured. Add it to .env and restart the app.")
    elif ai_provider == "OpenAI" and not OPENAI_API_KEY:
        st.warning("OpenAI is selected but OPENAI_API_KEY is not configured. Add it to .env and restart the app.")
    elif ai_provider == "Groq" and not GROQ_API_KEY:
        st.warning("Groq is selected but GROQ_API_KEY is not configured. Add it to .env and restart the app.")
    elif ai_provider == "Ollama (Local)" and not HAS_OLLAMA:
        st.warning("Ollama is selected but the ollama package is not installed. Run: pip install ollama")
    else:
        st.caption("OpenAI key: OPENAI_API_KEY. Gemini key: GEMINI_API_KEY or GOOGLE_API_KEY. Groq key: GROQ_API_KEY.")

# with st.container(border=True):
#     card_title("📂", "Dataset")
#     st.caption("Using local dataset: synthetic_transactions_complex_INTERNAL.csv")

input_df = None
try:
    input_df = load_data()
except Exception as ex:
    st.error(f"Unable to read dataset: {ex}")

if input_df is not None:
    missing = [c for c in SCHEMA_COLUMNS if c not in input_df.columns]
    if missing:
        st.error(f"Dataset does not match required schema. Missing columns: {missing}")
        st.stop()

    customer_lookup_df = pd.concat(
        [
            input_df[["originator_id", "originator_name"]].rename(
                columns={"originator_id": "customer_id", "originator_name": "customer_name"}
            ),
            input_df[["beneficiary_id", "beneficiary_name"]].rename(
                columns={"beneficiary_id": "customer_id", "beneficiary_name": "customer_name"}
            ),
        ],
        ignore_index=True,
    )
    customer_lookup_df["customer_id"] = customer_lookup_df["customer_id"].astype(str).str.strip()
    customer_lookup_df["customer_name"] = (
        customer_lookup_df["customer_name"].fillna("").astype(str).str.strip().replace({"nan": "", "None": ""})
    )
    customer_lookup_df = customer_lookup_df[customer_lookup_df["customer_id"] != ""].copy()

    id_name_map = {}
    for r in customer_lookup_df.itertuples(index=False):
        if r.customer_id not in id_name_map and r.customer_name:
            id_name_map[r.customer_id] = r.customer_name

    ids = sorted(
        set(input_df["originator_id"].astype(str).dropna().tolist()).union(
            set(input_df["beneficiary_id"].astype(str).dropna().tolist())
        )
    )

    default_target = "NGID-A2SDFC8IO3" if "NGID-A2SDFC8IO3" in ids else (ids[0] if ids else "")
    target_customer_id = st.selectbox(
        "Select Customer",
        ids,
        index=ids.index(default_target) if default_target in ids else 0,
        format_func=lambda cid: format_customer_label(str(cid), id_name_map.get(str(cid), "")),
    )

    run = st.button("Run analysis", type="primary")

    if run and int(n_hops) > 2:
        unique_node_count = estimate_unique_nodes_within_hops(
            df=input_df,
            target_customer_id=str(target_customer_id),
            use_all_dates=bool(use_all_dates),
            lookback_days=int(lookback_days),
            n_hops=int(n_hops),
        )
        st.session_state["pending_high_hop_analysis"] = {
            "target_customer_id": str(target_customer_id),
            "use_all_dates": bool(use_all_dates),
            "lookback_days": int(lookback_days),
            "n_hops": int(n_hops),
            "unique_node_count": int(unique_node_count),
        }
        run = False

    pending_high_hop = st.session_state.get("pending_high_hop_analysis")
    if st.session_state.pop("high_hop_cancel_notice", False):
        st.info("Analysis cancelled.")

    if pending_high_hop:
        pending_hops = int(pending_high_hop["n_hops"])
        pending_nodes = int(pending_high_hop["unique_node_count"])
        if pending_nodes > 50:
            high_hop_message = (
                f"Selected  n_hops={pending_hops} resulting total unique nodes={pending_nodes}, "
                "it will be time and rescource consuming, do you want to proceed ?"
            )
        else:
            high_hop_message = (
                f"Selected  n_hops={pending_hops} resulting total unique nodes={pending_nodes},  "
                "though the total unique nodes count is less than 50, but still it can be time and rescource consuming, do you want to proceed ?"
            )
        st.warning(
            high_hop_message
        )
        confirm_col, cancel_col = st.columns(2)
        with confirm_col:
            proceed_high_hop = st.button("Yes, proceed", key="confirm_high_hop_run")
        with cancel_col:
            cancel_high_hop = st.button("No, cancel", key="cancel_high_hop_run")

        if proceed_high_hop:
            run = True
            target_customer_id = str(pending_high_hop["target_customer_id"])
            use_all_dates = bool(pending_high_hop["use_all_dates"])
            lookback_days = int(pending_high_hop["lookback_days"])
            n_hops = int(pending_high_hop["n_hops"])
            st.session_state.pop("pending_high_hop_analysis", None)
        elif cancel_high_hop:
            st.session_state.pop("pending_high_hop_analysis", None)
            st.session_state["high_hop_cancel_notice"] = True
            st.rerun()

    if run:
        cfg = AnalyzerConfig(
            use_all_dates=bool(use_all_dates),
            lookback_days=int(lookback_days),
            n_hops=int(n_hops),
            ranking_weight_mode=parse_mode(ranking_mode),
        )

        analyzer = None
        outputs = None
        lookback_adjustment_message = None

        with st.spinner("Running network analysis..."):
            try:
                analyzer = FinancialCrimeNetworkAnalyzer(
                    df=input_df,
                    target_customer_id=target_customer_id,
                    config=cfg,
                )
                outputs = analyzer.run()
            except ValueError as ex:
                if (
                    str(ex) == "target_customer_id is not present in lookback-filtered transactions."
                    and not bool(use_all_dates)
                ):
                    dt_series = pd.to_datetime(input_df["transaction_datetime"], errors="coerce", utc=True).dt.tz_localize(None)
                    valid_mask = dt_series.notna()
                    if valid_mask.any():
                        valid_df = input_df.loc[valid_mask].copy()
                        valid_df["_txn_dt"] = dt_series.loc[valid_mask].values
                        cid = str(target_customer_id)
                        target_mask = (
                            (valid_df["originator_id"].astype(str) == cid)
                            | (valid_df["beneficiary_id"].astype(str) == cid)
                        )
                        target_tx = valid_df.loc[target_mask]

                        if not target_tx.empty:
                            max_dt = valid_df["_txn_dt"].max()
                            target_last_dt = target_tx["_txn_dt"].max()
                            min_required_lookback = max(int((max_dt - target_last_dt).days) + 1, 1)
                            adjusted_lookback = max(int(lookback_days), int(min_required_lookback))

                            adjusted_cfg = AnalyzerConfig(
                                use_all_dates=False,
                                lookback_days=int(adjusted_lookback),
                                n_hops=int(n_hops),
                                ranking_weight_mode=parse_mode(ranking_mode),
                            )

                            analyzer = FinancialCrimeNetworkAnalyzer(
                                df=input_df,
                                target_customer_id=target_customer_id,
                                config=adjusted_cfg,
                            )
                            outputs = analyzer.run()

                            lookback_adjustment_message = (
                                f"Requested lookback ({int(lookback_days)} days) did not include any transactions for selected customer "
                                f"{cid}. Latest transaction for this customer is on {target_last_dt.date()}, while dataset max date is "
                                f"{max_dt.date()}. Lookback was auto-adjusted to {int(adjusted_lookback)} days (minimum needed: "
                                f"{int(min_required_lookback)} days) so customer transactions can be analyzed."
                            )
                        else:
                            st.error(
                                f"Selected customer {cid} has no valid-dated transactions in the dataset, so lookback cannot be auto-adjusted."
                            )
                    else:
                        st.error("No valid transaction dates found in dataset; unable to adjust lookback automatically.")
                else:
                    st.error(f"Analysis failed: {ex}")
            except Exception as ex:
                st.error(f"Analysis failed: {ex}")

        if analyzer is not None and outputs is not None:
            st.session_state["analyzer"] = analyzer
            st.session_state["outputs"] = outputs
            st.session_state["target_customer_id"] = target_customer_id
            if lookback_adjustment_message:
                st.session_state["lookback_adjustment_message"] = lookback_adjustment_message
            else:
                st.session_state.pop("lookback_adjustment_message", None)

if "analyzer" in st.session_state:
    analyzer: FinancialCrimeNetworkAnalyzer = st.session_state["analyzer"]
    outputs = st.session_state["outputs"]
    target_customer_id = st.session_state["target_customer_id"]

    st.success("✅ Analysis completed")
    if st.session_state.get("lookback_adjustment_message"):
        st.info(st.session_state["lookback_adjustment_message"])

    # ── Part A: Executive Summary ──────────────────────────────────
    with st.container(border=True):
        card_title("📋", "Executive Summary")
        render_executive_summary_cards(analyzer, outputs)
        # Placeholder container for the target customer AI summary block under Executive Summary.
        ai_summary_container = st.container()

    # ── Network selector ──────────────────────────────────────────
    network_ids = analyzer.get_network_ids()
    if not network_ids:
        st.warning("No networks available for visualization.")
        st.stop()

    # ── Network Graph card ────────────────────────────────────────
    selected_node = ""
    with st.container(border=True):
        graph_title = "Full Network"
        card_title("🕸️", f"Network Graph — {graph_title}")
        network_focus = st.selectbox(
            "Network Focus",
            [analyzer.full_network_id] + network_ids,
            index=0,
            format_func=lambda n: "Full Network" if str(n) == str(analyzer.full_network_id) else str(n),
        )
        # Network Focus is the canonical selector for downstream network-specific views.
        selected_network = str(network_focus)
        if len(network_ids) > 1:
            st.caption("Select a focused split network to keep it active; other split networks are greyed out in the full graph.")

        node_df, edge_df = analyzer.get_network_graph_data(analyzer.full_network_id)
        cluster_map = analyzer.get_full_network_clusters() if len(network_ids) > 1 else None
        active_cluster_id = None
        if len(network_ids) > 1 and str(network_focus) != str(analyzer.full_network_id):
            try:
                active_cluster_id = int(str(network_focus).split("-")[-1])
            except Exception:
                active_cluster_id = None

        if node_df.empty:
            st.info("No graph data for selected network.")
        else:
            node_id_list = node_df["customer_id"].astype(str).tolist()
            node_name_map = (
                node_df[["customer_id", "customer_name"]]
                .drop_duplicates(subset=["customer_id"])
                .assign(customer_id=lambda x: x["customer_id"].astype(str))
                .set_index("customer_id")["customer_name"]
                .fillna("")
                .astype(str)
                .to_dict()
            )
            # Default to target customer, then session state, then first node
            default_idx = 0
            if str(target_customer_id) in node_id_list:
                default_idx = node_id_list.index(str(target_customer_id))
            elif "_graph_selected_node" in st.session_state and st.session_state["_graph_selected_node"] in node_id_list:
                default_idx = node_id_list.index(str(st.session_state["_graph_selected_node"]))

            if HAS_PLOTLY:
                fig = build_plotly_network_figure(
                    node_df,
                    edge_df,
                    target_customer_id,
                    cluster_map=cluster_map,
                    active_cluster_id=active_cluster_id,
                )
                if fig is not None:
                    try:
                        event = st.plotly_chart(
                            fig, use_container_width=True,
                            config={"scrollZoom": True},
                            on_select="rerun",
                            selection_mode="points",
                        )
                        if event and hasattr(event, "selection") and event.selection and event.selection.points:
                            clicked_id = str(event.selection.points[0].get("customdata", ""))
                            if clicked_id in node_id_list:
                                st.session_state["_graph_selected_node"] = clicked_id
                                default_idx = node_id_list.index(clicked_id)
                    except Exception:
                        st.plotly_chart(fig, use_container_width=True, config={"scrollZoom": True})
                    render_color_legend()
            elif HAS_AGRAPH:
                render_color_legend()
                graph_nodes, graph_edges = build_graph_objects(node_df, edge_df, target_customer_id)
                agraph_result = agraph(
                    nodes=graph_nodes,
                    edges=graph_edges,
                    config=Config(
                        width="100%", height=580, directed=True, physics=True,
                        hierarchical=False, nodeHighlightBehavior=True,
                        highlightColor="#f7a7a6", collapsible=False,
                    ),
                )
                if agraph_result and str(agraph_result) in node_id_list:
                    st.session_state["_graph_selected_node"] = str(agraph_result)
                    default_idx = node_id_list.index(str(agraph_result))
            else:
                render_color_legend()
                dot_graph = build_graphviz_dot(node_df, edge_df, target_customer_id)
                st.graphviz_chart(dot_graph, use_container_width=True)

            if len(network_ids) > 1:
                st.caption("Soft transparent highlights indicate split sub-networks inside the full 2-hop network.")

            with st.expander("📋 All nodes in this network", expanded=False):
                outward_metrics = (
                    edge_df.groupby("originator_id", as_index=False)
                    .agg(
                        outward_flow_volume=("txn_count", "sum"),
                        outward_flow_value_usd=("total_amount_usd", "sum"),
                        counterparties_outward=("beneficiary_id", "nunique"),
                    )
                    .rename(columns={"originator_id": "customer_id"})
                ) if not edge_df.empty else pd.DataFrame(columns=["customer_id", "outward_flow_volume", "outward_flow_value_usd", "counterparties_outward"])

                inward_metrics = (
                    edge_df.groupby("beneficiary_id", as_index=False)
                    .agg(
                        inward_flow_volume=("txn_count", "sum"),
                        inward_flow_value_usd=("total_amount_usd", "sum"),
                        counterparties_inward=("originator_id", "nunique"),
                    )
                    .rename(columns={"beneficiary_id": "customer_id"})
                ) if not edge_df.empty else pd.DataFrame(columns=["customer_id", "inward_flow_volume", "inward_flow_value_usd", "counterparties_inward"])

                all_nodes_df = node_df[[
                    "customer_id", "customer_name", "country", "entity_type",
                    "final_node_risk_score", "sanctions_flag", "pep_flag", "sar_flag", "exited_flag",
                ]].copy()

                if "network_id" in node_df.columns:
                    all_nodes_df["network_id"] = node_df["network_id"].astype(str)
                elif isinstance(cluster_map, dict) and cluster_map:
                    fallback_network_id = (
                        str(network_focus)
                        if str(network_focus) != str(analyzer.full_network_id)
                        else str(analyzer.full_network_id)
                    )
                    all_nodes_df["network_id"] = all_nodes_df["customer_id"].astype(str).map(
                        lambda cid: f"Network-{int(cluster_map[cid])}" if cid in cluster_map else fallback_network_id
                    )
                else:
                    all_nodes_df["network_id"] = (
                        str(network_focus)
                        if str(network_focus)
                        else str(analyzer.full_network_id)
                    )

                all_nodes_df = all_nodes_df.merge(outward_metrics, on="customer_id", how="left")
                all_nodes_df = all_nodes_df.merge(inward_metrics, on="customer_id", how="left")

                numeric_cols = [
                    "outward_flow_volume",
                    "outward_flow_value_usd",
                    "counterparties_outward",
                    "inward_flow_volume",
                    "inward_flow_value_usd",
                    "counterparties_inward",
                ]
                for c in numeric_cols:
                    all_nodes_df[c] = pd.to_numeric(all_nodes_df[c], errors="coerce").fillna(0)

                for c in ["outward_flow_volume", "counterparties_outward", "inward_flow_volume", "counterparties_inward"]:
                    all_nodes_df[c] = all_nodes_df[c].astype(int)

                display_df = all_nodes_df.rename(
                    columns={
                        "network_id": "Network ID",
                        "customer_id": "Customer ID",
                        "customer_name": "Customer Name",
                        "country": "Country",
                        "entity_type": "Entity Type",
                        "final_node_risk_score": "Overall Risk Score",
                        "sanctions_flag": "Sanctions Flag",
                        "pep_flag": "PEP Flag",
                        "sar_flag": "SAR Flag",
                        "exited_flag": "Exited Flag",
                        "inward_flow_value_usd": "Inward Flow Value (USD)",
                        "inward_flow_volume": "Inward Flow Volume",
                        "outward_flow_value_usd": "Outward Flow Value (USD)",
                        "outward_flow_volume": "Outward Flow Volume",
                        "counterparties_inward": "# Counterparties (Inward)",
                        "counterparties_outward": "# Counterparties (Outward)",
                    }
                )

                if "Network ID" in display_df.columns:
                    display_cols = ["Network ID"] + [c for c in display_df.columns if c != "Network ID"]
                    display_df = display_df[display_cols]

                st.dataframe(
                    display_df.sort_values("Overall Risk Score", ascending=False),
                    use_container_width=True,
                )

            selected_node = st.selectbox(
                "🔎 Click a node above or select here to load the case snapshot",
                node_id_list,
                index=default_idx,
                format_func=lambda cid: format_customer_label(str(cid), node_name_map.get(str(cid), "")),
            )
            st.session_state["_graph_selected_node"] = selected_node

            # Requested: hide score methodology/demo computation sections from UI.

    # ── Part B Row 1: Network summary ─────────────────────────────
    with st.container(border=True):
        card_title("📊", "Network Summary & Ranking")
        network_summary = outputs["network_rankings"].copy()
        st.dataframe(network_summary, use_container_width=True)

    # ── Part B Row 2: two side-by-side cards ──────────────────────
    col_left, col_right = st.columns(2)

    with col_left:
        with st.container(border=True):
            card_title("⚠️", "Top Risky Nodes")
            top_nodes_table = outputs["node_rankings"][
                [
                    "network_id",
                    "customer_id",
                    "customer_name",
                    "country",
                    "entity_type",
                    "final_node_risk_score",
                    "key_reasons",
                ]
            ].copy()
            top_nodes_table = top_nodes_table.sort_values(["network_id", "final_node_risk_score"], ascending=[True, False])
            top_nodes_table = top_nodes_table.groupby("network_id").head(20).reset_index(drop=True)
            st.dataframe(top_nodes_table, use_container_width=True)

    with col_right:
        with st.container(border=True):
            card_title("🚨", "Theme / Subtheme Trigger Log")
            theme_table = outputs["theme_triggers"][
                ["network_id", "theme", "subtheme", "severity_score", "example_transaction_ids"]
            ].copy()
            theme_table = theme_table.sort_values(["network_id", "severity_score"], ascending=[True, False])
            st.dataframe(theme_table, use_container_width=True)

    # ── Case Snapshot ─────────────────────────────────────────────
    ai_summary = ""
    ai_actions = analyzer.build_next_actions()
    ai_status = ""
    target_ai_summary = ""
    target_ai_actions = analyzer.build_next_actions()
    target_ai_status = ""

    if selected_node:
        # Keep linked transactions aligned with the graph scope (Network Focus) and selected node.
        txn_scope_network = str(selected_network)
        node_rankings = outputs.get("node_rankings", pd.DataFrame())
        effective_selected_network = resolve_effective_network_for_node(
            analyzer=analyzer,
            node_rankings=node_rankings,
            selected_network=selected_network,
            selected_node=str(selected_node),
            target_customer_id=str(target_customer_id),
        )
        effective_target_network = resolve_effective_network_for_node(
            analyzer=analyzer,
            node_rankings=node_rankings,
            selected_network=selected_network,
            selected_node=str(target_customer_id),
            target_customer_id=str(target_customer_id),
        )
        node_tx = analyzer.get_node_transactions(txn_scope_network, selected_node)
        kyc = analyzer.get_customer_kyc(selected_node)
        target_tx = analyzer.get_node_transactions(effective_target_network, target_customer_id)
        target_kyc = analyzer.get_customer_kyc(target_customer_id)

        all_network_frames = []
        for nid, net_df in analyzer.network_txn.items():
            export_df = net_df.copy()
            export_df.insert(0, "network_id", str(nid))
            all_network_frames.append(export_df)
        all_networks_tx_export = (
            pd.concat(all_network_frames, ignore_index=True)
            if all_network_frames
            else pd.DataFrame(columns=["network_id"] + list(analyzer.df.columns))
        )

        if str(txn_scope_network) == str(analyzer.full_network_id):
            scoped_network_tx_export = analyzer.full_network_txn.copy()
            if not scoped_network_tx_export.empty:
                if not all_networks_tx_export.empty and "transaction_id" in scoped_network_tx_export.columns:
                    txn_to_network = all_networks_tx_export[["transaction_id", "network_id"]].drop_duplicates("transaction_id")
                    scoped_network_tx_export = scoped_network_tx_export.merge(
                        txn_to_network,
                        on="transaction_id",
                        how="left",
                    )
                    scoped_network_tx_export["network_id"] = scoped_network_tx_export["network_id"].fillna(str(analyzer.full_network_id))
                    cols = ["network_id"] + [c for c in scoped_network_tx_export.columns if c != "network_id"]
                    scoped_network_tx_export = scoped_network_tx_export[cols]
                else:
                    scoped_network_tx_export.insert(0, "network_id", str(analyzer.full_network_id))
        else:
            scoped_network_tx_export = analyzer.network_txn.get(str(txn_scope_network), pd.DataFrame()).copy()
            if scoped_network_tx_export.empty:
                scoped_network_tx_export = pd.DataFrame(columns=["network_id"] + list(analyzer.df.columns))
            else:
                scoped_network_tx_export.insert(0, "network_id", str(txn_scope_network))

        target_ai_summary, target_ai_actions, target_context = build_investigator_target_summary(
            analyzer=analyzer,
            outputs=outputs,
            selected_network=selected_network,
            target_customer_id=str(target_customer_id),
            node_df=node_df,
            target_tx=target_tx,
            kyc=target_kyc,
        )
        target_ai_status = "AI Summary Generator deterministic summary generated from local analyzer outputs."

        if use_ai_agent and ai_provider == "OpenAI":
            if OPENAI_API_KEY:
                try:
                    with st.spinner(f"Generating target customer investigator summary with OpenAI ({ai_model})..."):
                        model_target_summary = _generate_openai_investigator_summary(
                            target_context=target_context,
                            baseline_text=target_ai_summary,
                            model=ai_model,
                        )
                        model_target_actions = _generate_openai_next_actions(
                            node_context=target_context,
                            fallback_actions=target_ai_actions,
                            model=ai_model,
                        )
                    if model_target_summary.strip():
                        target_ai_summary = model_target_summary
                    if model_target_actions:
                        target_ai_actions = model_target_actions
                    target_ai_status = f"AI Summary Generator used OpenAI model '{ai_model}' for the target customer summary."
                except Exception as ex:
                    target_ai_status = _openai_error_status(ex)
            else:
                target_ai_status = "OPENAI_API_KEY is not configured; showing deterministic AI Summary Generator output."
        elif use_ai_agent and ai_provider == "Gemini":
            if GEMINI_API_KEY:
                try:
                    with st.spinner(f"Generating target customer investigator summary with Gemini ({ai_model})..."):
                        model_target_summary = _generate_gemini_investigator_summary(
                            target_context=target_context,
                            baseline_text=target_ai_summary,
                            model=ai_model,
                        )
                        model_target_actions = _generate_gemini_next_actions(
                            node_context=target_context,
                            fallback_actions=target_ai_actions,
                            model=ai_model,
                        )
                    if model_target_summary.strip():
                        target_ai_summary = model_target_summary
                    if model_target_actions:
                        target_ai_actions = model_target_actions
                    target_ai_status = f"AI Summary Generator used Gemini model '{ai_model}' for the target customer summary."
                except Exception as ex:
                    target_ai_status = f"Gemini unavailable for target customer summary ({ex}); showing deterministic AI Summary Generator output."
            else:
                target_ai_status = "GEMINI_API_KEY is not configured; showing deterministic AI Summary Generator output."
        elif use_ai_agent and ai_provider == "Groq":
            if GROQ_API_KEY:
                try:
                    with st.spinner(f"Generating target customer investigator summary with Groq ({ai_model})..."):
                        model_target_summary = _generate_groq_investigator_summary(
                            target_context=target_context,
                            baseline_text=target_ai_summary,
                            model=ai_model,
                        )
                        model_target_actions = _generate_groq_next_actions(
                            node_context=target_context,
                            fallback_actions=target_ai_actions,
                            model=ai_model,
                        )
                    if model_target_summary.strip():
                        target_ai_summary = model_target_summary
                    if model_target_actions:
                        target_ai_actions = model_target_actions
                    target_ai_status = f"AI Summary Generator used Groq model '{ai_model}' for the target customer summary."
                except Exception as ex:
                    target_ai_status = f"Groq unavailable for target customer summary ({ex}); showing deterministic AI Summary Generator output."
            else:
                target_ai_status = "GROQ_API_KEY is not configured; showing deterministic AI Summary Generator output."
        elif use_ai_agent and ai_provider == "Ollama (Local)" and HAS_OLLAMA:
            try:
                with st.spinner(f"Generating target customer investigator summary with Ollama ({ollama_model})..."):
                    model_target_summary, model_target_actions = generate_ollama_investigator_summary(
                        model_name=ollama_model.strip() or "llama3.1:8b",
                        target_customer_id=str(target_context["target_customer_id"]),
                        customer_name=str(target_context["customer_name"]),
                        selected_network=str(target_context["selected_network"]),
                        window_label=str(target_context["window_label"]),
                        network_count=int(target_context["network_count"]),
                        network_score=float(target_context["network_score"]),
                        node_score=float(target_context["node_score"]),
                        risk_band=str(target_context["risk_band"]),
                        txn_count=int(target_context["txn_count"]),
                        total_value=float(target_context["total_value"]),
                        outgoing_count=int(target_context["outgoing_count"]),
                        incoming_count=int(target_context["incoming_count"]),
                        top_driver_text=str(target_context["top_driver_text"]),
                        flag_text=target_context["flag_text"],
                        top_theme_text=str(target_context["top_theme_text"]),
                        top_counterparty_text=str(target_context["top_counterparty_text"]),
                        top_scenario_text=str(target_context["top_scenario_text"]),
                    )
                if model_target_summary.strip():
                    target_ai_summary = model_target_summary
                if model_target_actions:
                    target_ai_actions = model_target_actions
                target_ai_status = f"AI Summary Generator used Ollama model '{(ollama_model.strip() or 'llama3.1:8b')}' for the target customer summary."
            except Exception as ex:
                target_ai_status = f"Ollama unavailable for target customer summary ({ex}); showing deterministic AI Summary Generator output."
        elif not use_ai_agent or ai_provider == "Deterministic":
            target_ai_status = "AI Summary Generator disabled; showing deterministic summary/actions."
        elif ai_provider == "Ollama (Local)" and not HAS_OLLAMA:
            target_ai_status = "Ollama Python package not installed; showing deterministic AI Summary Generator output."
        else:
            target_ai_status = "AI summary provider not configured; showing deterministic AI Summary Generator output."

        ai_actions = target_ai_actions

        with ai_summary_container:
            card_title("🤖", f"AI Summary Generator — Investigator Summary for Target Customer {target_customer_id}")
            st.markdown(f'<div class="genai-card">{target_ai_summary.replace(chr(10), "<br>")}</div>', unsafe_allow_html=True)
            if target_ai_status:
                st.caption(target_ai_status)
            st.download_button(
                "📥 Download Target Customer Summary",
                target_ai_summary.encode("utf-8"),
                f"target_customer_{target_customer_id}_investigator_summary.txt",
                "text/plain",
            )

        fallback_summary = build_selected_customer_genai_summary(
            selected_network=effective_selected_network,
            selected_customer_id=str(selected_node),
            node_df=node_df,
            node_rankings=outputs["node_rankings"],
            node_tx=node_tx,
        )
        rank_row = node_rankings[
            (node_rankings["network_id"].astype(str) == str(effective_selected_network))
            & (node_rankings["customer_id"].astype(str) == str(selected_node))
        ]
        rank = rank_row.iloc[0] if not rank_row.empty else None

        network_row_df = analyzer.network_summary[
            analyzer.network_summary["network_id"].astype(str) == str(effective_selected_network)
        ]
        network_row = network_row_df.iloc[0] if not network_row_df.empty else None

        top_themes_df = analyzer.theme_log[
            analyzer.theme_log["network_id"].astype(str) == str(effective_selected_network)
        ].sort_values("severity_score", ascending=False).head(5)
        theme_names = [str(r.subtheme) for r in top_themes_df.itertuples(index=False)]

        txn_count = int(len(node_tx))
        total_value = float(node_tx["amount_usd"].sum()) if txn_count else 0.0
        avg_value = float(node_tx["amount_usd"].mean()) if txn_count else 0.0
        outgoing_count = int((node_tx["originator_id"].astype(str) == str(selected_node)).sum()) if txn_count else 0
        incoming_count = int((node_tx["beneficiary_id"].astype(str) == str(selected_node)).sum()) if txn_count else 0

        sanctions = bool(rank["sanctions_flag"]) if rank is not None and "sanctions_flag" in rank else False
        pep = bool(rank["pep_flag"]) if rank is not None and "pep_flag" in rank else False
        sar = bool(rank["sar_flag"]) if rank is not None and "sar_flag" in rank else False
        exited = bool(rank["exited_flag"]) if rank is not None and "exited_flag" in rank else False
        node_score = float(rank.get("final_node_risk_score", 0.0) or 0.0) if rank is not None else 0.0
        network_score = float(network_row.get("network_risk_score", 0.0) or 0.0) if network_row is not None else 0.0

        pagerank_c = float(rank.get("pagerank_component", 0.0) or 0.0) if rank is not None else 0.0
        flags_c = float(rank.get("flags_component", 0.0) or 0.0) if rank is not None else 0.0
        prox_c = float(rank.get("proximity_component", 0.0) or 0.0) if rank is not None else 0.0
        beh_c = float(rank.get("behaviour_component", 0.0) or 0.0) if rank is not None else 0.0

        drivers = [
            ("pagerank influence", pagerank_c),
            ("direct risk flags", flags_c),
            ("proximity to flagged entities", prox_c),
            ("behavioural patterns", beh_c),
        ]
        drivers = sorted(drivers, key=lambda x: x[1], reverse=True)
        top_driver_text = ", ".join([f"{k} ({v:.1f})" for k, v in drivers[:3]])

        flag_text = []
        if sanctions:
            flag_text.append("Sanctions")
        if pep:
            flag_text.append("PEP")
        if sar:
            flag_text.append("SAR")
        if exited:
            flag_text.append("Exited")

        node_context = {
            "selected_node": str(selected_node),
            "selected_network": str(effective_selected_network),
            "node_score": node_score,
            "network_score": network_score,
            "txn_count": txn_count,
            "total_value": total_value,
            "avg_value": avg_value,
            "outgoing_count": outgoing_count,
            "incoming_count": incoming_count,
            "top_driver_text": top_driver_text,
            "flag_text": flag_text,
            "top_theme_text": ", ".join(theme_names[:4]) if theme_names else "none",
            "top_counterparty_text": "none",
        }

        deterministic_summary, deterministic_actions, _ = build_ai_customer_brief(
            analyzer=analyzer,
            outputs=outputs,
            selected_network=effective_selected_network,
            selected_node=str(selected_node),
            node_df=node_df,
            node_tx=node_tx,
            fallback_summary=fallback_summary,
            fallback_actions=ai_actions,
        )
        ai_summary = deterministic_summary
        ai_actions = deterministic_actions
        ai_status = (
            "Deterministic summary shown by default. Click 'Generate AI Summary for Selected Node' to use the selected AI provider."
            if use_ai_agent
            else "AI agent disabled; showing deterministic summary/actions."
        )

        current_node_key = f"{selected_network}|{effective_selected_network}|{selected_node}"
        previous_node_key = st.session_state.get("_selected_node_ai_key")
        if previous_node_key != current_node_key:
            st.session_state["_selected_node_ai_key"] = current_node_key
            st.session_state.pop("selected_node_ai_cached", None)

        cached_node_ai = st.session_state.get("selected_node_ai_cached")
        if isinstance(cached_node_ai, dict) and cached_node_ai.get("node_key") == current_node_key:
            ai_summary = str(cached_node_ai.get("summary", ai_summary))
            cached_actions = cached_node_ai.get("actions", ai_actions)
            ai_actions = cached_actions if isinstance(cached_actions, list) else ai_actions
            ai_status = str(cached_node_ai.get("status", ai_status))

        with st.container(border=True):
            card_title("🗂️", f"Case Snapshot — {selected_node}")
            render_case_snapshot(
                analyzer=analyzer,
                selected_node=selected_node,
                target_customer_id=str(target_customer_id),
                selected_network=selected_network,
                full_network_id=str(analyzer.full_network_id),
                txn_scope_network=txn_scope_network,
                node_df=node_df,
                node_rankings=outputs["node_rankings"],
                node_tx=node_tx,
                kyc=kyc,
                scoped_network_tx_export=scoped_network_tx_export,
                all_networks_tx_export=all_networks_tx_export,
            )

        with st.container(border=True):
            card_title("🤖", "Selected Node AI Risk Summary")
            generate_node_ai_clicked = False
            if use_ai_agent:
                generate_node_ai_clicked = st.button(
                    "Generate AI Summary for Selected Node",
                    key="selected_node_ai_generate_btn",
                )

            if generate_node_ai_clicked:
                if ai_provider == "OpenAI":
                    if OPENAI_API_KEY:
                        try:
                            with st.spinner(f"Generating selected node summary with OpenAI ({ai_model})..."):
                                model_summary = _generate_openai_node_risk_summary(
                                    node_context=node_context,
                                    baseline_text=fallback_summary,
                                    model=ai_model,
                                )
                                model_actions = _generate_openai_next_actions(
                                    node_context=node_context,
                                    fallback_actions=ai_actions,
                                    model=ai_model,
                                )
                            ai_summary = model_summary if model_summary.strip() else fallback_summary
                            ai_actions = model_actions if model_actions else ai_actions
                            ai_status = f"OpenAI summary generated using model '{ai_model}'."
                        except Exception as ex:
                            ai_summary = fallback_summary
                            ai_status = _openai_error_status(ex)
                    else:
                        ai_summary = fallback_summary
                        ai_status = "OPENAI_API_KEY is not configured; showing deterministic summary/actions."
                elif ai_provider == "Gemini":
                    if GEMINI_API_KEY:
                        try:
                            with st.spinner(f"Generating selected node summary with Gemini ({ai_model})..."):
                                model_summary = _generate_gemini_node_risk_summary(
                                    node_context=node_context,
                                    baseline_text=fallback_summary,
                                    model=ai_model,
                                )
                                model_actions = _generate_gemini_next_actions(
                                    node_context=node_context,
                                    fallback_actions=ai_actions,
                                    model=ai_model,
                                )
                            ai_summary = model_summary if model_summary.strip() else fallback_summary
                            ai_actions = model_actions if model_actions else ai_actions
                            ai_status = f"Gemini summary generated using model '{ai_model}'."
                        except Exception as ex:
                            ai_summary = fallback_summary
                            ai_status = f"Gemini unavailable ({ex}); showing deterministic summary/actions."
                    else:
                        ai_summary = fallback_summary
                        ai_status = "GEMINI_API_KEY is not configured; showing deterministic summary/actions."
                elif ai_provider == "Groq":
                    if GROQ_API_KEY:
                        try:
                            with st.spinner(f"Generating selected node summary with Groq ({ai_model})..."):
                                model_summary = _generate_groq_node_risk_summary(
                                    node_context=node_context,
                                    baseline_text=fallback_summary,
                                    model=ai_model,
                                )
                                model_actions = _generate_groq_next_actions(
                                    node_context=node_context,
                                    fallback_actions=ai_actions,
                                    model=ai_model,
                                )
                            ai_summary = model_summary if model_summary.strip() else fallback_summary
                            ai_actions = model_actions if model_actions else ai_actions
                            ai_status = f"Groq summary generated using model '{ai_model}'."
                        except Exception as ex:
                            ai_summary = fallback_summary
                            ai_status = f"Groq unavailable ({ex}); showing deterministic summary/actions."
                    else:
                        ai_summary = fallback_summary
                        ai_status = "GROQ_API_KEY is not configured; showing deterministic summary/actions."
                elif ai_provider == "Ollama (Local)":
                    if not HAS_OLLAMA:
                        ai_summary = fallback_summary
                        ai_status = "Ollama Python package not installed; showing deterministic summary/actions."
                    else:
                        try:
                            with st.spinner(f"Generating summary with Ollama ({ollama_model})..."):
                                model_summary, model_actions = generate_ollama_customer_brief(
                                    model_name=ollama_model.strip() or "llama3.1:8b",
                                    selected_network=str(selected_network),
                                    selected_node=str(selected_node),
                                    node_score=node_score,
                                    network_score=network_score,
                                    txn_count=txn_count,
                                    total_value=total_value,
                                    avg_value=avg_value,
                                    outgoing_count=outgoing_count,
                                    incoming_count=incoming_count,
                                    top_driver_text=top_driver_text,
                                    flag_text=flag_text,
                                    theme_names=theme_names,
                                )
                            ai_summary = model_summary if model_summary.strip() else fallback_summary
                            ai_actions = model_actions if model_actions else ai_actions
                            ai_status = f"Ollama summary generated using model '{(ollama_model.strip() or 'llama3.1:8b')}'."
                        except Exception as ex:
                            ai_summary = fallback_summary
                            ai_status = f"Ollama unavailable ({ex}); showing deterministic summary/actions."
                else:
                    ai_summary = deterministic_summary
                    ai_actions = deterministic_actions
                    ai_status = "Deterministic provider selected; showing deterministic summary/actions."

                st.session_state["selected_node_ai_cached"] = {
                    "node_key": current_node_key,
                    "summary": ai_summary,
                    "actions": ai_actions,
                    "status": ai_status,
                }

            st.markdown(f'<div class="genai-card">{ai_summary.replace(chr(10), "<br>")}</div>', unsafe_allow_html=True)
            if ai_status:
                st.caption(ai_status)

    # ── AI Model Recommendation card ──────────────────────────────
    with st.container(border=True):
        _rec_n_hops = int(getattr(analyzer.config, "n_hops", 2))
        _rec_lookback_days = int(getattr(analyzer.config, "lookback_days", 180))
        card_title(
            "🧭",
            f"AI Model Recommendation — {_rec_n_hops}-Hop Network Risk Assessment "
            f"({_rec_lookback_days}-day lookback)",
        )
        st.caption(
            f"Scans networks up to {_rec_n_hops} hops from the target for SAR, Sanctions, PEP, Exit, or DRA-alert "
            f"flags within the last {_rec_lookback_days} days, then recommends strict action with connected-customer "
            "detail and rationale, or Enhanced Due Diligence if no significant risk is found. "
            f"Uses the AI provider selected in the sidebar ({ai_provider}). Hop radius and lookback window match "
            "the values set in the sidebar for the current analysis run."
        )
        if st.button("Generate AI Model Recommendation", key="ai_network_recommendation_btn"):
            with st.spinner(f"Running Financial Crime Network Analyst prompt with {ai_provider}..."):
                recommendation_text, recommendation_status = generate_ai_network_recommendation(
                    analyzer=analyzer,
                    outputs=outputs,
                    target_customer_id=str(target_customer_id),
                    ai_provider=ai_provider,
                    ai_model=ai_model,
                    ollama_model=ollama_model,
                )
            st.session_state["ai_network_recommendation_text"] = recommendation_text
            st.session_state["ai_network_recommendation_status"] = recommendation_status

        if "ai_network_recommendation_text" in st.session_state:
            st.markdown(
                f'<div class="genai-card">{st.session_state["ai_network_recommendation_text"].replace(chr(10), "<br>")}</div>',
                unsafe_allow_html=True,
            )
            st.caption(st.session_state["ai_network_recommendation_status"])
            st.download_button(
                "📥 Download AI Model Recommendation",
                st.session_state["ai_network_recommendation_text"].encode("utf-8"),
                f"ai_model_recommendation_{target_customer_id}.txt",
                "text/plain",
                key="ai_network_recommendation_download_btn",
            )

    export_recommendation_text = str(st.session_state.get("ai_network_recommendation_text", "") or "")
    export_recommendation_status = str(st.session_state.get("ai_network_recommendation_status", "") or "")
    if not export_recommendation_text:
        export_context_text = _build_network_flag_context(analyzer, outputs, str(target_customer_id))
        export_recommendation_text = _build_deterministic_network_recommendation(
            export_context_text,
            n_hops=int(getattr(analyzer.config, "n_hops", 2)),
            lookback_days=int(getattr(analyzer.config, "lookback_days", 180)),
        )
        export_recommendation_status = (
            "AI model recommendation not generated in the UI; included deterministic network recommendation in the PDF export."
        )

    selected_node_for_export = str(selected_node) if "selected_node" in locals() else str(target_customer_id)
    network_summary_pdf_bytes = build_network_summary_pdf_bytes(
        analyzer=analyzer,
        target_customer_id=str(target_customer_id),
        selected_network=str(selected_network),
        selected_node=selected_node_for_export,
        network_summary_df=network_summary,
        top_nodes_df=top_nodes_table,
        theme_table_df=theme_table,
        target_summary_text=target_ai_summary,
        target_summary_status=target_ai_status,
        target_actions=target_ai_actions,
        selected_node_summary_text=ai_summary,
        selected_node_summary_status=ai_status,
        selected_node_actions=ai_actions,
        recommendation_text=export_recommendation_text,
        recommendation_status=export_recommendation_status,
    )
    network_summary_excel_bytes = build_network_summary_excel_bytes(
        analyzer=analyzer,
        target_customer_id=str(target_customer_id),
        selected_network=str(selected_network),
        selected_node=selected_node_for_export,
        network_summary_df=network_summary,
        top_nodes_df=top_nodes_table,
        theme_table_df=theme_table,
        all_networks_tx_df=all_networks_tx_export,
        target_summary_text=target_ai_summary,
        target_summary_status=target_ai_status,
        target_actions=target_ai_actions,
        selected_node_summary_text=ai_summary,
        selected_node_summary_status=ai_status,
        selected_node_actions=ai_actions,
        recommendation_text=export_recommendation_text,
        recommendation_status=export_recommendation_status,
    )

    # ── Downloads card ────────────────────────────────────────────
    with st.container(border=True):
        card_title("⬇️", "Download Results")
        c1, c2, _spacer = st.columns([1, 1, 3])
        with c1:
            st.download_button(
                "📥 Full Summary Excel",
                network_summary_excel_bytes,
                f"network_summary_{target_customer_id}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_network_summary_excel_btn",
            )
        with c2:
            st.download_button(
                "📥 Full Summary PDF",
                network_summary_pdf_bytes,
                f"network_summary_{target_customer_id}.pdf",
                "application/pdf",
                key="download_network_summary_pdf_btn",
            )

    with st.container(border=True):
        card_title("💬", "Ask Atlas")
        st.text_area(
            "Ask Atlas",
            value="Ask anything to Atlas regarding the networks found",
            disabled=True,
            label_visibility="collapsed",
            height=90,
        )

else:
    st.info("Select target customer and click Run analysis.")
